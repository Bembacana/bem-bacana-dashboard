#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bem Bacana — Dashboard Financeiro v2.0
Consolida Planilha2 (2023-dez/2025) + Eloca (2026+) e gera output/dashboard.html interativo.

Execução:
    cd scripts && python3 gerar_dashboard.py
"""

import os, sys, json, base64
from datetime import datetime
from collections import defaultdict
import openpyxl

# ─── CONFIGURAÇÃO ─────────────────────────────────────────────────────────────
META_MENSAL = 150_000.0
MESES_PT    = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

VERT_MAP_P2 = {
    'EVENTOS':            'Eventos',
    'CASA':               'Casa',
    'ESCRITÓRIO':         'Escritório',
    'ESCRITORIO':         'Escritório',
    'CONCIERGE':          'Concierge',
    'SEMINOVOS':          'Seminovos',
    'COMISSÃO':           'Outras Receitas',
    'COMISSAO':           'Outras Receitas',
    'REPOSIÇÃO/AVARIAS':  'Outras Receitas',
    'REPOSICAO/AVARIAS':  'Outras Receitas',
}
VERT_MAP_ELOCA = {
    'LOCAÇÃO EVENTOS':    'Eventos',
    'LOCACAO EVENTOS':    'Eventos',
    'LOCAÇÃO CASA':       'Casa',
    'LOCACAO CASA':       'Casa',
    'LOCAÇÃO ESCRITÓRIO': 'Escritório',
    'LOCACAO ESCRITORIO': 'Escritório',
    'VENDA DE SEMINOVOS': 'Seminovos',
}
VERT_MAP_ERP = {
    'Vertical Eventos':    'Eventos',
    'VERTICAL EVENTOS':    'Eventos',
    'Vertical Casa':       'Casa',
    'VERTICAL CASA':       'Casa',
    'Vertical Escritório': 'Escritório',
    'Vertical Escritórios':'Escritório',
    'CONCIERGE':           'Concierge',
    'SEMINOVOS':           'Seminovos',
    'REPOSIÇÃO/AVARIAS':   'Outras Receitas',
}
STATUS_MAP_ERP = {
    'Fechada':    'Aprovada',
    'Cancelada':  'Cancelada',
    'Aguardando': 'Em negociação',
}
STATUS_MAP_ELOCA = {
    'PROPOSTA APROVADA':            'Aprovada',
    'PROPOSTA FINALIZADA':          'Aprovada',
    'PROPOSTA RENOVADA (CASA)':     'Aprovada',
    'PROPOSTA ENVIADA':             'Enviada',
    'EM NEGOCIAÇÃO & AJUSTES':      'Em negociação',
    'EM NEGOCIACAO & AJUSTES':      'Em negociação',
    'EM ORÇAMENTO':                 'Em negociação',
    'PROPOSTA REPROVADA':           'Reprovada',
    'CLIENTE NÃO DEU CONTINUIDADE': 'Sem continuidade',
    'CLIENTE NAO DEU CONTINUIDADE': 'Sem continuidade',
    'CANCELADO':                    'Cancelada',
}
VERTICAIS_PRINCIPAIS = ['Eventos', 'Casa', 'Escritório', 'Concierge', 'Seminovos']
COR_VERTICAL = {
    'Eventos':         '#076A76',
    'Casa':            '#41A8B9',
    'Escritório':      '#2E86AB',
    'Concierge':       '#FBAE4B',
    'Seminovos':       '#B0865A',
    'Outras Receitas': '#A9A69F',
}
COR_STATUS = {
    'Aprovada':         '#076A76',
    'Enviada':          '#41A8B9',
    'Em negociação':    '#FBAE4B',
    'Reprovada':        '#DC3545',
    'Sem continuidade': '#A9A69F',
    'Cancelada':        '#61605B',
    'Preliminar':      '#9B59B6',
}

# ─── LOGO ─────────────────────────────────────────────────────────────────────
def carregar_logo(base_dir):
    for p in [
        os.path.join(base_dir, 'dados', 'logo_bb.png'),
        os.path.join(base_dir, '..', 'dados', 'logo_bb.png'),
    ]:
        if os.path.exists(p):
            with open(p, 'rb') as f:
                return base64.b64encode(f.read()).decode()
    return None

# ─── LEITURA ──────────────────────────────────────────────────────────────────
def ler_planilha2(path):
    """Faturamento histórico 2023 – dez/2025."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Planilha2']
    registros, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None]*8
        data, cliente, valor, vertical, cli_novo, tipo, dur, pag = row[:8]
        if not data or not isinstance(data, datetime) or not valor:
            continue
        if data > datetime(2025, 12, 31):   # Planilha2 cobre até dez/2025
            continue
        v    = str(vertical or '').strip().upper()
        vert = VERT_MAP_P2.get(v, 'Outras Receitas')
        novo = cli_novo is True or str(cli_novo or '').strip().upper() in ('TRUE','VERDADAIRO','VERDADEIRO','1')
        pago = str(pag or '').strip() != 'Inadimplente'
        val_f = float(valor)
        key  = (str(cliente or '').strip(), data.date(), vert, round(val_f, 2))
        if key in seen:
            continue
        seen.add(key)
        registros.append({
            'fonte': 'P2',
            'data': data, 'ano': data.year, 'mes': data.month,
            'ym': f'{data.year}-{data.month:02d}',
            'cliente': str(cliente or '').strip(),
            'valor': val_f, 'vertical': vert,
            'novo': novo, 'pago': pago,
        })
    wb.close()
    return registros


def _parse_date_eloca(v):
    for fmt in ('%d/%m/%Y','%Y-%m-%d'):
        try: return datetime.strptime(str(v).strip(), fmt)
        except: pass
    return None

def ler_eloca(path):
    """Pipeline e faturamento 2026 em diante. Suporta 3 formatos por deteccao de headers."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Eloca']
    # Detectar formato pelo cabeçalho
    headers = [str(c or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
    has_contrato = 'contrato' in headers
    has_contato  = 'nome do contato' in headers  # Relatorio (15): 11 cols
    has_desconto = 'val. desconto' in headers
    has_frete    = 'val. frete' in headers        # Relatorio (13): tem frete separado
    has_status   = 'status' in headers            # Relatorio (19)+: coluna Status extra entre Tipo e Fase
    n_cols       = len(headers)
    if has_contato:
        # Relatorio (15): 11 cols com Nome do Contato
        # Proposta(0), Cliente(1), Contrato(2), Contato(3), Tipo(4), Fase(5), DataCad(6), DataIni(7), DataFim(8), ValServ(9), ValProp(10)
        IX = dict(id=0,cli=1,ctr=2,tipo=4,fase=5,dcad=6,dini=7,dfim=8,vserv=9,vprop=10)
    elif has_contrato and has_desconto and has_frete:
        # Relatorio (13): 12 cols com Val.Frete + Val.Desconto + Val.Proposta
        # Proposta(0)..DataFim(7), ValServ(8), ValFrete(9), ValDesconto(10), ValProp(11)
        IX = dict(id=0,cli=1,ctr=2,tipo=3,fase=4,dcad=5,dini=6,dfim=7,vserv=8,vprop=11)
    elif has_contrato and has_desconto and has_status:
        # Relatorio (19)+: 12 cols, Status(4) inserido entre Tipo e Fase
        # Proposta(0), Cliente(1), Contrato(2), Tipo(3), Status(4), Fase(5), DataCad(6), DataIni(7), DataFim(8), ValServ(9), ValDesconto(10), ValProp(11)
        IX = dict(id=0,cli=1,ctr=2,tipo=3,fase=5,dcad=6,dini=7,dfim=8,vserv=9,vprop=11)
    elif has_contrato and has_desconto:
        # Relatorio (18): 12 cols SEM Val.Frete — ValServ(8), ValDesconto(9), ValProp(10)
        IX = dict(id=0,cli=1,ctr=2,tipo=3,fase=4,dcad=5,dini=6,dfim=7,vserv=8,vprop=10)
    elif has_contrato and n_cols <= 10:
        # Relatorio (16): 10 cols — ValServ(8), ValProp(9)
        IX = dict(id=0,cli=1,ctr=2,tipo=3,fase=4,dcad=5,dini=6,dfim=7,vserv=8,vprop=9)
    elif has_contrato:
        # Generico com contrato >10 cols sem desconto
        IX = dict(id=0,cli=1,ctr=2,tipo=3,fase=4,dcad=5,dini=6,dfim=7,vserv=8,vprop=9)
    else:
        # 9 cols sem contrato: Proposta, Cliente, Tipo, Fase, DataCad, DataInicio, DataFim, ValServ, ValProp
        IX = dict(id=0,cli=1,ctr=None,tipo=2,fase=3,dcad=4,dini=5,dfim=6,vserv=7,vprop=8)

    registros, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None]*15
        id_p = row[IX['id']]
        if not id_p or id_p in seen:
            continue
        seen.add(id_p)
        dt_cad  = _parse_date_eloca(row[IX['dcad']])
        if not dt_cad:
            continue
        dt_ini = _parse_date_eloca(row[IX['dini']]) if row[IX['dini']] else None
        dt_fim = _parse_date_eloca(row[IX['dfim']]) if row[IX['dfim']] else None
        contrato = row[IX['ctr']] if IX['ctr'] is not None else None
        tipo_v   = str(row[IX['tipo']] or '').strip().upper()
        vert     = VERT_MAP_ELOCA.get(tipo_v, 'Outras Receitas')
        fase_s   = str(row[IX['fase']] or '').strip()
        status   = STATUS_MAP_ELOCA.get(fase_s, 'Outros')
        renovacao = fase_s == 'PROPOSTA RENOVADA (CASA)'
        vs = float(row[IX['vserv']] or 0)
        vp = float(row[IX['vprop']] or 0)
        duracao = (dt_fim - dt_ini).days if dt_ini and dt_fim else None
        antecedencia = (dt_ini - dt_cad).days if dt_ini else None
        registros.append({
            'fonte': 'Eloca',
            'id': id_p, 'contrato': contrato,
            'data': dt_cad, 'ano': dt_cad.year, 'mes': dt_cad.month,
            'ym': f'{dt_cad.year}-{dt_cad.month:02d}',
            'data_inicio': dt_ini, 'data_fim': dt_fim,
            'duracao_dias': duracao, 'antecedencia_dias': antecedencia,
            'cliente': str(row[IX['cli']] or '').strip(),
            'vertical': vert, 'status': status,
            'fase_original': fase_s, 'renovacao': renovacao,
            'valor_total':   vp,
            'valor_locacao': round(vp - vs, 2),
            'valor_frete':   vs,
        })
    wb.close()
    return registros


def ler_erp_crm(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['ERP_CRM']
    registros, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None]*8
        num, cliente, tipo, total, status, dt_cri = row[:6]
        if not num or num in seen:
            continue
        seen.add(num)
        if not isinstance(dt_cri, datetime):
            continue
        if dt_cri > datetime(2025, 12, 31):
            continue
        vert    = VERT_MAP_ERP.get(str(tipo or '').strip(), 'Outras Receitas')
        status_n = STATUS_MAP_ERP.get(str(status or '').strip(), str(status or ''))
        registros.append({
            'fonte': 'ERP_CRM', 'id': num,
            'data': dt_cri, 'ano': dt_cri.year, 'mes': dt_cri.month,
            'ym': f'{dt_cri.year}-{dt_cri.month:02d}',
            'cliente': str(cliente or '').strip(),
            'vertical': vert, 'status': status_n,
            'valor_total': float(total or 0),
        })
    wb.close()
    return registros



def ler_crm(path):
    """Lê planilha de Propostas Preliminares (CRM externo ao ERP).
    Formato: 9 cols — Proposta, Nome do Cliente, Tipo, Fase Negociação,
    Data Cadastro, Data Início, Data Fim, Val. Serviços, Val. Proposta."""
    FASES_ERP = {
        'PROPOSTA APROVADA', 'PROPOSTA FINALIZADA', 'PROPOSTA RENOVADA (CASA)'
    }
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    registros, seen = [], set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None]*10
        id_p  = row[0]
        cli   = row[1]
        tipo_v = str(row[2] or '').strip().upper()
        fase_s = str(row[3] or '').strip()
        dcad_s = row[4]
        dini_s = row[5]
        dfim_s = row[6]
        vs    = row[7]
        vp    = row[8]
        if not id_p or id_p in seen:
            continue
        seen.add(id_p)
        dt_cad = _parse_date_eloca(dcad_s)
        if not dt_cad:
            continue
        dt_ini = _parse_date_eloca(dini_s) if dini_s else None
        dt_fim = _parse_date_eloca(dfim_s) if dfim_s else None
        vert   = VERT_MAP_ELOCA.get(tipo_v, 'Outras Receitas')
        status = STATUS_MAP_ELOCA.get(fase_s, 'Outros')
        vp_f   = float(vp or 0)
        vs_f   = float(vs or 0)
        em_erp = fase_s.upper() in {f.upper() for f in FASES_ERP}
        registros.append({
            'id': id_p,
            'cliente': str(cli or '').strip(),
            'data': dt_cad,
            'ano': dt_cad.year,
            'mes': dt_cad.month,
            'ym': f'{dt_cad.year}-{dt_cad.month:02d}',
            'data_inicio': dt_ini,
            'data_fim': dt_fim,
            'vertical': vert,
            'status': status,
            'fase_original': fase_s,
            'valor_total': vp_f,
            'valor_locacao': round(vp_f - vs_f, 2),
            'em_erp': em_erp,
        })
    wb.close()
    return registros

def ler_cac(path):
    MESES_CAC = {
        'Janeiro':1,'Fevereiro':2,'Março':3,'Abril':4,
        'Maio':5,'Junho':6,'Julho':7,'Agosto':8,
        'Setembro':9,'Outubro':10,'Novembro':11,'Dezembro':12,
    }
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Plan CAC']
    cac = {}
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        if not row[1] or str(row[1]).strip() not in MESES_CAC:
            continue
        m = MESES_CAC[str(row[1]).strip()]
        cac[f'2024-{m:02d}'] = {
            'novos': int(row[2] or 0),
            'investimento': float(row[3] or 0),
            'cac': float(row[4] or 0),
        }
    wb.close()
    return cac

def ler_faturamento(path):
    """Lê planilha de faturas emitidas pelo financeiro. Detecta formato pelo header."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Detectar layout pelo cabeçalho
    hdr = [str(c or '').strip().lower() for c in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
    # Formatos suportados:
    # 13 cols (novo, "Planilha de Faturamento DD.MM.AAAA"):
    #   Fatura, Emissão, Faturamento, Proposta, Contrato, Tipo, Vencimento, Razão Social,
    #   Valor Total Fatura, %, Rep 2, Tipo Receita, Valor Líquido
    # 10 cols (com município): Faturamento, Proposta, Contrato, Tipo, Cliente, Razão Social, Município, UF, Valor Total Fatura, Tipo Receita
    # 9 cols: Faturamento, Proposta, Contrato, Tipo, Cliente, Razão Social, Valor Total Fatura, Tipo Receita, Valor Líquido
    # 7 cols antigo: Faturamento, Proposta, Contrato, Cliente, Razão Social, Valor Total Fatura, Tipo Receita
    has_fatura_col = hdr[0] == 'fatura' if hdr else False  # 13-col: primeira col é "Fatura" (nº da fatura)
    has_emissao    = any('emiss' in h for h in hdr)
    has_tipo       = 'tipo' in hdr and hdr.index('tipo') == 3
    has_municipio  = any('munic' in h for h in hdr)
    if has_fatura_col and has_emissao:
        # 13 cols: Fatura(0), Emissão(1), Faturamento(2), Proposta(3), Contrato(4), Tipo(5),
        #          Vencimento(6), Razão Social(7), Valor Total Fatura(8), %(9), Rep2(10),
        #          Tipo Receita(11), Valor Líquido(12)
        IX_dt=1; IX_razao=7; IX_valor=8; IX_tr=11
    elif has_tipo and has_municipio:
        # 10 cols: Faturamento, Proposta, Contrato, Tipo, Cliente, Razão Social, Município, UF, Valor Total Fatura, Tipo Receita
        IX_dt=0; IX_razao=5; IX_valor=8; IX_tr=9
    elif has_tipo:
        # 9 cols: Faturamento, Proposta, Contrato, Tipo, Cliente, Razão Social, Valor Total Fatura, Tipo Receita, ...
        IX_dt=0; IX_razao=5; IX_valor=6; IX_tr=7
    else:
        # 7 cols antigo: Faturamento, Proposta, Contrato, Cliente, Razão Social, Valor Total Fatura, Tipo Receita
        IX_dt=0; IX_razao=4; IX_valor=5; IX_tr=6
    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None]*8
        dt_str = row[IX_dt]
        razao  = row[IX_razao]
        valor  = row[IX_valor]
        tipo_receita = row[IX_tr]
        if not dt_str or not valor:
            continue
        try:
            dt = datetime.strptime(str(dt_str).strip(), '%d/%m/%Y')
        except:
            continue
        tipo_raw = str(tipo_receita or '').strip()
        # Simplificar: "LOCAÇÃO CASA PJ [1193]" → "Casa"
        tipo_simples = 'Outros'
        tr_up = tipo_raw.upper()
        has_pj = 'PJ' in tr_up
        if 'CASA' in tr_up:
            tipo_simples = 'Casa - PJ' if has_pj else 'Casa - PF'
        elif 'EVENTOS' in tr_up:
            tipo_simples = 'Eventos - PJ' if has_pj else 'Eventos - PF'
        elif 'ESCRITÓRIO' in tr_up or 'ESCRITORIO' in tr_up: tipo_simples = 'Escritório'
        elif 'CONCIERGE' in tr_up: tipo_simples = 'Concierge'
        elif 'SEMINOVOS' in tr_up: tipo_simples = 'Seminovos'
        registros.append({
            'data': dt, 'ano': dt.year, 'mes': dt.month,
            'ym': f'{dt.year}-{dt.month:02d}',
            'cliente': str(razao or '').strip(),
            'valor': float(str(valor).replace(',','.')),
            'tipo_receita': tipo_raw,
            'tipo_simples': tipo_simples,
        })
    wb.close()
    return registros


def ler_manual_input(dados_dir):
    path = os.path.join(dados_dir, 'manual_input.json')
    default = {
        'marketing': {}, 'inadimplencia': {},
        'inadimplencia_v2': [],
        'propostas_preliminares': {},
        'faturamento_manual': {},
    }
    if not os.path.exists(path):
        return default
    with open(path, 'r', encoding='utf-8') as f:
        d = json.load(f)
    for k in default:
        if k not in d:
            d[k] = default[k]
    return d


# ─── CÁLCULO DE KPIs ──────────────────────────────────────────────────────────
def calcular(p2, eloca, erp, cac_hist, faturamento=None, manual=None, crm=None):
    # ── Faturamento e clientes ──
    fat       = defaultdict(lambda: defaultdict(float))
    fat_br    = defaultdict(lambda: defaultdict(float))
    cli_m     = defaultdict(set)
    novos_m   = defaultdict(int)
    rec_m     = defaultdict(int)

    for r in p2:
        if not r['pago']:
            continue
        fat[r['ym']][r['vertical']] += r['valor']
        cli_m[r['ym']].add(r['cliente'])

    for r in eloca:
        if r['status'] != 'Aprovada':
            continue
        if r['ano'] < 2026:  # Planilha2 já cobre até dez/2025 — evita dupla contagem
            continue
        fat[r['ym']][r['vertical']] += r['valor_total']
        fat_br[r['ym']]['locacao']  += r['valor_locacao']
        fat_br[r['ym']]['frete']    += r['valor_frete']
        cli_m[r['ym']].add(r['cliente'])

    # ── Novo vs Recorrente ERP: passagem cronológica por nome normalizado ───────
    def _norm(s): return ' '.join(str(s or '').upper().split())
    cli_m_norm    = {ym: {_norm(c) for c in cset} for ym, cset in cli_m.items()}
    cli_m_names   = {ym: {_norm(c): c for c in cset} for ym, cset in cli_m.items()}
    # Mapa de vertical por cliente normalizado por mês (cards de Evolução)
    cli_vert_m = defaultdict(dict)   # {ym: {norm_name: vertical}}
    for r in p2:
        if not r['pago']: continue
        cli_vert_m[r['ym']][_norm(r['cliente'])] = r['vertical']
    for r in eloca:
        if r['status'] != 'Aprovada' or r['ano'] < 2026: continue
        cli_vert_m[r['ym']][_norm(r['cliente'])] = r['vertical']
    vistos_global = set()
    erp_detalhes  = {}  # {ym: {novos:[{nome,vertical}], antigos:[{nome,vertical}]}}
    for ym in sorted(cli_m_norm.keys()):
        cset_norm = cli_m_norm[ym]
        novos_set  = cset_norm - vistos_global
        antigos_set = cset_norm & vistos_global
        novos_m[ym] = len(novos_set)
        rec_m[ym]   = len(antigos_set)
        name_map    = cli_m_names[ym]
        vmap        = cli_vert_m.get(ym, {})
        erp_detalhes[ym] = {
            'novos':   sorted([{'nome': name_map.get(n,n), 'vertical': vmap.get(n,'')} for n in novos_set], key=lambda x: x['nome']),
            'antigos': sorted([{'nome': name_map.get(n,n), 'vertical': vmap.get(n,'')} for n in antigos_set], key=lambda x: x['nome']),
        }
        vistos_global |= cset_norm
    # Conjunto global de todos os clientes ERP (para cruzamento com CRM)
    erp_all_norm = vistos_global.copy()

    # ── Novo vs Recorrente CRM: cruzamento com base ERP ─────────────────────
    crm_cli_m  = defaultdict(dict)  # {ym: {nome_norm: nome_original}}
    crm_vert_m = defaultdict(dict)  # {ym: {nome_norm: vertical}}
    for r in (crm or []):
        n = _norm(r['cliente'])
        if n:
            crm_cli_m[r['ym']][n]  = r['cliente']
            crm_vert_m[r['ym']][n] = r.get('vertical', '')
    crm_detalhes = {}
    for ym, cmap in crm_cli_m.items():
        novos_crm   = {n: orig for n, orig in cmap.items() if n not in erp_all_norm}
        antigos_crm = {n: orig for n, orig in cmap.items() if n in erp_all_norm}
        vmap_c = crm_vert_m.get(ym, {})
        crm_detalhes[ym] = {
            'novos':   sorted([{'nome': orig, 'vertical': vmap_c.get(n,'')} for n, orig in novos_crm.items()], key=lambda x: x['nome']),
            'antigos': sorted([{'nome': orig, 'vertical': vmap_c.get(n,'')} for n, orig in antigos_crm.items()], key=lambda x: x['nome']),
        }

    # ── Pipeline ──
    pip_st      = defaultdict(lambda: {'qtd':0,'valor':0.0})
    pip_sub     = defaultdict(lambda: {'qtd':0,'valor':0.0})
    pip_mes     = defaultdict(lambda: defaultdict(int))
    pip_mes_val = defaultdict(lambda: defaultdict(float))   # valor por status por mês
    pip_sub_mes = defaultdict(lambda: defaultdict(lambda: {'qtd':0,'valor':0.0}))  # sub-tipo por mês

    for r in erp:
        pip_st[r['status']]['qtd']    += 1
        pip_st[r['status']]['valor']  += r['valor_total']
        pip_mes[r['ym']][r['status']] += 1
        pip_mes_val[r['ym']][r['status']] += r['valor_total']

    for r in eloca:
        pip_st[r['status']]['qtd']    += 1
        pip_st[r['status']]['valor']  += r['valor_total']
        pip_mes[r['ym']][r['status']] += 1
        pip_mes_val[r['ym']][r['status']] += r['valor_total']
        pip_sub[r['fase_original']]['qtd']   += 1
        pip_sub[r['fase_original']]['valor'] += r['valor_total']
        pip_sub_mes[r['ym']][r['fase_original']]['qtd']   += 1
        pip_sub_mes[r['ym']][r['fase_original']]['valor'] += r['valor_total']

    # ── Taxa de conversão (fórmula corrigida) ──
    aprov_qtd = pip_st.get('Aprovada',{}).get('qtd',0)
    neg_qtd   = (pip_st.get('Cancelada',{}).get('qtd',0)
               + pip_st.get('Reprovada',{}).get('qtd',0)
               + pip_st.get('Sem continuidade',{}).get('qtd',0))
    denom_qtd = aprov_qtd + neg_qtd
    taxa_qtd  = round(aprov_qtd / denom_qtd * 100, 1) if denom_qtd else 0.0

    aprov_val  = pip_st.get('Aprovada',{}).get('valor',0)
    # Denominador = valor total de TODAS as propostas enviadas (todos os status)
    denom_val  = sum(v.get('valor',0) for v in pip_st.values())
    taxa_val   = round(aprov_val / denom_val * 100, 1) if denom_val else 0.0
    # Valor absoluto aprovado (para exibir em R$ no card)
    aprov_val_abs = aprov_val

    # ── Série temporal ──
    # Adicionar meses de execução futura (data_inicio Eloca) para visibilidade de 2027+
    exec_yms = {r['data_inicio'].strftime('%Y-%m') for r in eloca
                if r.get('data_inicio') and r['data_inicio'].year >= 2026}
    todos_ym = sorted(set(fat.keys()) | set(pip_mes.keys()) | exec_yms)
    serie = []
    for i, ym in enumerate(todos_ym):
        ano, mes = int(ym[:4]), int(ym[5:])
        total    = sum(fat[ym].values())
        prev     = sum(fat[todos_ym[i-1]].values()) if i > 0 else 0
        prev_yr  = sum(fat.get(f'{ano-1}-{mes:02d}', {}).values())
        mom = round((total - prev)    / prev    * 100, 1) if prev    else None
        yoy = round((total - prev_yr) / prev_yr * 100, 1) if prev_yr else None
        serie.append({
            'ym': ym, 'ano': ano, 'mes': mes,
            'label': f"{MESES_PT[mes-1]}/{str(ano)[2:]}",
            'faturamento':  round(total, 2),
            'fat_locacao':  round(fat_br[ym].get('locacao',0), 2),
            'fat_frete':    round(fat_br[ym].get('frete',0), 2),
            'por_vertical': {v: round(fat[ym].get(v,0),2) for v in VERTICAIS_PRINCIPAIS+['Outras Receitas']},
            'meta':     META_MENSAL,
            'pct_meta': round(total / META_MENSAL * 100, 1),
            'mom': mom, 'yoy': yoy,
            'clientes':    len(cli_m.get(ym, set())),
            'novos':       novos_m.get(ym, 0),
            'recorrentes': rec_m.get(ym, 0),
            'cac_hist':    cac_hist.get(ym,{}).get('cac'),
            'inv_mkt_hist':cac_hist.get(ym,{}).get('investimento')
                           or (manual or {}).get('marketing',{}).get(ym)
                           or None,
            'pipeline':    dict(pip_mes.get(ym,{})),
            'pipeline_val':dict(pip_mes_val.get(ym,{})),
            'pip_sub':     {k:{'qtd':v['qtd'],'valor':v['valor']} for k,v in pip_sub_mes.get(ym,{}).items()},
        })

    # ── Top clientes (all-time + por mês) ──
    top_all = defaultdict(lambda: {'valor':0.0,'contratos':0})
    for r in p2:
        if r['pago'] and r['cliente']:
            top_all[r['cliente']]['valor']     += r['valor']
            top_all[r['cliente']]['contratos'] += 1
    for r in eloca:
        if r['status']=='Aprovada' and r['cliente'] and r['ano'] >= 2026:
            top_all[r['cliente']]['valor']     += r['valor_total']
            top_all[r['cliente']]['contratos'] += 1
    top_list = [{'nome':k,'valor':round(v['valor'],2),'contratos':v['contratos']}
                for k,v in sorted(top_all.items(), key=lambda x:-x[1]['valor'])[:20]]

    top_por_mes = {}
    for ym in todos_ym:
        tm = defaultdict(lambda: {'valor':0.0,'contratos':0})
        for r in p2:
            if r['pago'] and r['ym']==ym and r['cliente']:
                tm[r['cliente']]['valor']     += r['valor']
                tm[r['cliente']]['contratos'] += 1
        for r in eloca:
            if r['status']=='Aprovada' and r['ym']==ym and r['cliente'] and r['ano'] >= 2026:
                tm[r['cliente']]['valor']     += r['valor_total']
                tm[r['cliente']]['contratos'] += 1
        top_por_mes[ym] = [{'nome':k,'valor':round(v['valor'],2),'contratos':v['contratos']}
                           for k,v in sorted(tm.items(),key=lambda x:-x[1]['valor'])[:10]]

    # ── Ticket médio por vertical (all-time + por mês) ──
    def _ticket(records_p2, records_eloca):
        out = {}
        for v in VERTICAIS_PRINCIPAIS:
            vals = [r['valor'] for r in records_p2 if r['vertical']==v and r['pago']]
            vals += [r['valor_total'] for r in records_eloca
                     if r['vertical']==v and r['status']=='Aprovada' and r['ano'] >= 2026]
            out[v] = {
                'total':       round(sum(vals),2),
                'contratos':   len(vals),
                'ticket_medio':round(sum(vals)/len(vals),2) if vals else 0,
            }
        return out

    ticket_all = _ticket(p2, eloca)
    ticket_por_mes = {}
    for ym in todos_ym:
        p2_ym = [r for r in p2 if r['ym']==ym]
        el_ym = [r for r in eloca if r['ym']==ym]
        ticket_por_mes[ym] = _ticket(p2_ym, el_ym)

    # ── Frequência de clientes antigos (já alugaram antes) ──
    freq = defaultdict(lambda: {'locacoes':0,'valor':0.0,'verticais':defaultdict(int)})
    for r in p2:
        if r['pago'] and r['cliente']:
            freq[r['cliente']]['locacoes'] += 1
            freq[r['cliente']]['valor']    += r['valor']
            freq[r['cliente']]['verticais'][r['vertical']] += 1
    for r in eloca:
        if r['status']=='Aprovada' and r['cliente'] and r['ano'] >= 2026:
            freq[r['cliente']]['locacoes'] += 1
            freq[r['cliente']]['valor']    += r['valor_total']
            freq[r['cliente']]['verticais'][r['vertical']] += 1
    def _top_vert(vd): return max(vd, key=vd.get) if vd else '—'
    freq_list = [{'nome':k,'locacoes':v['locacoes'],'valor':round(v['valor'],2),
                  'vertical':_top_vert(v['verticais'])}
                 for k,v in sorted(freq.items(), key=lambda x:-x[1]['locacoes'])[:50]]

    # freq_por_mes: {ym: [{nome, locacoes, valor, vertical}]} para ranking filtrado por período
    freq_pm = defaultdict(lambda: defaultdict(lambda: {'locacoes':0,'valor':0.0,'verticais':defaultdict(int)}))
    for r in p2:
        if r['pago'] and r['cliente']:
            freq_pm[r['ym']][r['cliente']]['locacoes'] += 1
            freq_pm[r['ym']][r['cliente']]['valor']    += r['valor']
            freq_pm[r['ym']][r['cliente']]['verticais'][r['vertical']] += 1
    for r in eloca:
        if r['status']=='Aprovada' and r['cliente'] and r['ano'] >= 2026:
            freq_pm[r['ym']][r['cliente']]['locacoes'] += 1
            freq_pm[r['ym']][r['cliente']]['valor']    += r['valor_total']
            freq_pm[r['ym']][r['cliente']]['verticais'][r['vertical']] += 1
    freq_por_mes = {
        ym: [{'nome':k,'locacoes':v['locacoes'],'valor':round(v['valor'],2),'vertical':_top_vert(v['verticais'])}
             for k,v in clientes.items()]
        for ym, clientes in freq_pm.items()
    }

    # ── Mix receita (Eloca) ──
    mix_por_mes = {}
    for ym, vals in fat_br.items():
        mix_por_mes[ym] = {'locacao':round(vals.get('locacao',0),2),
                           'frete':  round(vals.get('frete',0),2)}

    # ── Faturamento por ano ──
    fat_ano = defaultdict(float)
    for r in p2:
        if r['pago']:
            fat_ano[r['ano']] += r['valor']
    for r in eloca:
        if r['status']=='Aprovada' and r['ano'] >= 2026:
            fat_ano[r['ano']] += r['valor_total']

    # ── Sub-tipos aprovação (Eloca) ──
    sub_keys = ['PROPOSTA APROVADA','PROPOSTA FINALIZADA','PROPOSTA RENOVADA (CASA)']
    aprov_sub = {k: {'qtd':pip_sub.get(k,{}).get('qtd',0),
                     'valor':round(pip_sub.get(k,{}).get('valor',0),2)}
                 for k in sub_keys}

    # ── Propostas preliminares (manual) ────────────────────────────────────────
    prelim = manual.get('propostas_preliminares', {}) if manual else {}
    # Adicionar ao pip_st como status 'Preliminar'
    for ym, entradas in prelim.items():
        for e in entradas:
            pip_st['Preliminar']['qtd']   += 1
            pip_st['Preliminar']['valor'] += float(e.get('valor', 0))
            pip_mes[ym]['Preliminar']     += 1
            pip_mes_val[ym]['Preliminar'] += float(e.get('valor', 0))

    # ── Macro propostas (todas as fases + preliminares) por mês ─────────────
    macro = {}
    for ym in sorted(set(pip_mes.keys())):
        qtd   = sum(pip_mes[ym].values())
        valor = sum(pip_mes_val[ym].values())
        macro[ym] = {'qtd': qtd, 'valor': round(valor, 2)}

    # ── Análise de Data Início / Data Fim ───────────────────────────────────
    from collections import Counter
    entregas_semana  = Counter()
    entregas_mes     = Counter()
    duracoes_casa    = defaultdict(list)
    duracoes_eventos = defaultdict(list)
    antecedencia_mes = defaultdict(list)

    for r in eloca:
        if not r.get('data_inicio'):
            continue
        di = r['data_inicio']
        yw = f"{di.isocalendar()[0]}-S{di.isocalendar()[1]:02d}"
        ym_ini = f'{di.year}-{di.month:02d}'
        entregas_semana[yw] += 1
        entregas_mes[ym_ini] += 1
        if r['duracao_dias'] is not None:
            vert = (r.get('vertical') or '').upper()
            if 'CASA' in vert:
                duracoes_casa[ym_ini].append(r['duracao_dias'])
            elif 'EVENTO' in vert:
                duracoes_eventos[ym_ini].append(r['duracao_dias'])
        if r['antecedencia_dias'] is not None and r['antecedencia_dias'] >= 0:
            antecedencia_mes[ym_ini].append(r['antecedencia_dias'])

    all_dur_casa    = [d for v in duracoes_casa.values()    for d in v]
    all_dur_eventos = [d for v in duracoes_eventos.values() for d in v]

    entregas_data = {
        'por_semana': {k: entregas_semana[k]
                       for k in sorted(entregas_semana)[-26:]},
        'por_mes':    {k: entregas_mes[k] for k in sorted(entregas_mes)},
        'duracao_casa':    {ym: round(sum(v)/len(v),1) for ym, v in duracoes_casa.items() if v},
        'duracao_eventos': {ym: round(sum(v)/len(v),1) for ym, v in duracoes_eventos.items() if v},
        'duracao_media_global': {
            'Casa':    round(sum(all_dur_casa)/len(all_dur_casa),1)       if all_dur_casa    else None,
            'Eventos': round(sum(all_dur_eventos)/len(all_dur_eventos),1) if all_dur_eventos else None,
        },
        'antecedencia_media': {ym: round(sum(v)/len(v),1)
                               for ym, v in antecedencia_mes.items() if v},
    }

    # ── Faturamento financeiro ───────────────────────────────────────────────
    fat_fin = defaultdict(lambda: defaultdict(float))
    fat_fin_total = defaultdict(float)
    if faturamento:
        for r in faturamento:
            fat_fin[r['ym']][r['tipo_simples']] += r['valor']
            fat_fin_total[r['ym']] += r['valor']
    # Manual
    for ym, entradas in (manual or {}).get('faturamento_manual', {}).items():
        for e in entradas:
            fat_fin[ym][e.get('tipo_receita','Outros')] += float(e.get('valor',0))
            fat_fin_total[ym] += float(e.get('valor',0))

    fat_fin_data = {
        'por_mes': {ym: {
            'total': round(fat_fin_total[ym], 2),
            'por_tipo': {k: round(v, 2) for k, v in fat_fin[ym].items()},
        } for ym in sorted(fat_fin_total.keys())},
    }


    # ── Pipeline CRM (Propostas Preliminares — fora do ERP) ─────────────────
    from collections import defaultdict as _dd
    pip_crm_mes     = _dd(lambda: _dd(lambda: {'qtd':0,'valor':0.0}))
    pip_crm_total   = _dd(lambda: {'qtd':0,'valor':0.0})
    crm_entregas_ab = []
    FASES_ABERTO_CRM = {'Enviada', 'Em negociação'}
    for r in (crm or []):
        if r['em_erp']:
            continue
        st = r['status']
        pip_crm_mes[r['ym']][st]['qtd']   += 1
        pip_crm_mes[r['ym']][st]['valor'] += r['valor_total']
        pip_crm_total[st]['qtd']          += 1
        pip_crm_total[st]['valor']        += r['valor_total']
        if st in FASES_ABERTO_CRM:
            crm_entregas_ab.append({
                'id':       r['id'],
                'cliente':  r['cliente'],
                'vertical': r['vertical'],
                'status':   st,
                'ym':       r['ym'],
                'data_inicio': r['data_inicio'].strftime('%d/%m/%Y') if r['data_inicio'] else '',
                'data_fim':   r['data_fim'].strftime('%d/%m/%Y')   if r['data_fim']   else '',
                'valor':    r['valor_total'],
            })
    crm_pipeline_mes = {
        ym: {st: {'qtd': v['qtd'], 'valor': round(v['valor'],2)}
             for st, v in statuses.items()}
        for ym, statuses in pip_crm_mes.items()
    }
    crm_pipeline_total = {
        st: {'qtd': v['qtd'], 'valor': round(v['valor'],2)}
        for st, v in pip_crm_total.items()
    }

    return {
        'serie':        serie,
        'pipeline':     {k:{'qtd':v['qtd'],'valor':round(v['valor'],2)} for k,v in pip_st.items()},
        'aprov_sub':    aprov_sub,
        'taxa_qtd':     taxa_qtd,
        'taxa_val':     taxa_val,
        'aprov_val_abs': round(aprov_val_abs, 2),
        'top_clientes': top_list,
        'top_por_mes':  top_por_mes,
        'ticket':       ticket_all,
        'ticket_por_mes': ticket_por_mes,
        'fat_ano':      {str(k):round(v,2) for k,v in sorted(fat_ano.items())},
        'freq_clientes':freq_list,
        'freq_por_mes': freq_por_mes,
        'mix_por_mes':  mix_por_mes,
        'meta_mensal':  META_MENSAL,
        'atualizado':   datetime.now().strftime('%d/%m/%Y %H:%M'),
        'macro':        macro,
        'entregas':     entregas_data,
        'fat_fin':      fat_fin_data,
        'inadimplencia_v2': (manual or {}).get('inadimplencia_v2', []),
        'prelim':       prelim,
        'erp_detalhes':  erp_detalhes,
        'crm_detalhes':  crm_detalhes,
        'crm_pipeline_mes':   crm_pipeline_mes,
        'crm_pipeline_total': crm_pipeline_total,
        'crm_entregas_aberto': crm_entregas_ab,
    }

# ─── GERAÇÃO DO HTML ──────────────────────────────────────────────────────────
def brl(v):
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')


# ─── GERAÇÃO DO HTML ──────────────────────────────────────────────────────────
def gerar_html(kpis, logo_b64):
    logo_src   = f'data:image/png;base64,{logo_b64}' if logo_b64 else ''
    data_json  = json.dumps(kpis, ensure_ascii=False)
    atualizado = kpis['atualizado']
    # Datas de atualização por fonte (exibidas nos cabeçalhos de seção)
    _fd        = kpis.get('fonte_datas', {})
    _erp_dt    = f' <span style="font-size:9px;font-weight:400;opacity:.7;margin-left:8px">atualizado {_fd["erp"]}</span>' if _fd.get('erp') else ''
    _crm_dt    = f' <span style="font-size:9px;font-weight:400;opacity:.7;margin-left:8px">atualizado {_fd["crm"]}</span>' if _fd.get('crm') else ''
    _fat_dt    = f' <span style="font-size:9px;font-weight:400;opacity:.7;margin-left:8px">atualizado {_fd["faturamento"]}</span>' if _fd.get('faturamento') else ''
    anos_disponiveis = sorted({s['ano'] for s in kpis['serie']})
    # Sempre garantir o próximo ano calendario (2027+) nos filtros
    from datetime import datetime as _dt
    _next_yr = _dt.now().year + 1
    if _next_yr not in anos_disponiveis:
        anos_disponiveis.append(_next_yr)
        anos_disponiveis.sort()
    anos_opts  = ''.join(f'<button class="btn-period" data-year="{a}">{a}</button>' for a in anos_disponiveis)
    logo_tag   = f'<img src="{logo_src}" alt="Bem Bacana" style="height:47px">' if logo_src else '<strong style="color:var(--primary);font-size:23px;letter-spacing:2px">BEM BACANA</strong>'
    cor_status_js = json.dumps({
        'Aprovada':'#076A76','Enviada':'#41A8B9','Em negociação':'#FBAE4B',
        'Reprovada':'#DC3545','Sem continuidade':'#A9A69F','Cancelada':'#61605B'})
    cor_vert_js = json.dumps({
        'Eventos':'#076A76','Casa':'#41A8B9','Escritório':'#2E86AB',
        'Concierge':'#FBAE4B','Seminovos':'#B0865A','Outras Receitas':'#A9A69F'})
    verts_js = json.dumps(['Eventos','Casa','Escritório','Concierge','Seminovos'])
    pip_order_js = json.dumps(['Aprovada','Enviada','Em negociação','Reprovada','Sem continuidade','Cancelada'])

    css = """
:root{--primary:#076A76;--secondary:#41A8B9;--accent:#FBAE4B;--text:#61605B;
      --white:#fff;--bg:#F4F6F7;--card:#fff;--border:#E8E6E1;--gray:#A9A69F;
      --green:#28A745;--red:#DC3545;--r:10px;--sh:0 1px 8px rgba(0,0,0,.07)}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}
.hdr{background:var(--primary);padding:0 32px;display:flex;justify-content:space-between;align-items:center;
     position:sticky;top:0;z-index:99;box-shadow:0 2px 16px rgba(7,106,118,.35);min-height:64px}
.hdr-logo{background:#fff;border-radius:8px;padding:7px 14px;display:flex;align-items:center}
.hdr-right{font-size:14px;color:rgba(255,255,255,.7);text-align:right;line-height:1.6}
.hdr-right strong{color:#fff;font-size:17px}
.filter-bar{background:#fff;border-bottom:1px solid var(--border);padding:10px 32px;
            display:flex;align-items:center;gap:10px;flex-wrap:wrap;
            position:sticky;top:64px;z-index:98;box-shadow:0 2px 6px rgba(0,0,0,.04)}
.filter-label{font-size:11px;font-weight:700;letter-spacing:1.2px;color:var(--gray);text-transform:uppercase;margin-right:4px}
.btn-period{border:1.5px solid var(--border);background:#fff;color:var(--text);border-radius:20px;
            padding:4px 14px;font-size:12px;font-family:inherit;cursor:pointer;transition:.15s}
.btn-period:hover{border-color:var(--primary);color:var(--primary)}
.btn-period.active{background:var(--primary);border-color:var(--primary);color:#fff;font-weight:600}
.month-select{border:1.5px solid var(--border);border-radius:20px;padding:4px 10px;
              font-size:12px;font-family:inherit;color:var(--text);background:#fff;cursor:pointer}
.month-select:focus{outline:none;border-color:var(--primary)}
#filter-desc{margin-left:auto;font-size:12px;color:var(--gray);font-style:italic}
#period-custom-form{display:none;align-items:center;gap:6px;margin-left:8px}
#period-custom-form.open{display:flex}
.period-input{border:1.5px solid var(--border);border-radius:6px;padding:3px 7px;font-size:12px;color:var(--text);background:#fff;height:28px}
.period-input:focus{outline:none;border-color:var(--primary)}
#btn-period-custom{border:1.5px solid var(--secondary);background:#fff;color:var(--secondary);border-radius:20px;padding:3px 11px;font-size:12px;cursor:pointer;font-weight:600}
#btn-period-custom.active{background:var(--secondary);color:#fff}
.period-apply{background:var(--primary);color:#fff;border:none;border-radius:6px;padding:3px 10px;font-size:12px;cursor:pointer}
.period-clear{background:#f1f3f5;color:var(--gray);border:none;border-radius:6px;padding:3px 8px;font-size:12px;cursor:pointer}
.wrap{max-width:1420px;margin:0 auto;padding:24px 28px 56px}
.sec-label{font-size:10px;font-weight:700;letter-spacing:1.8px;color:var(--gray);text-transform:uppercase;
           margin-bottom:10px;margin-top:26px}
.grid-kpi{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:14px;margin-bottom:6px}
.kpi{background:var(--card);border-radius:var(--r);padding:18px 20px;box-shadow:var(--sh)}
.kpi label{font-size:10px;font-weight:700;letter-spacing:1.2px;color:var(--gray);text-transform:uppercase;display:block;margin-bottom:6px}
.kpi .v{font-size:26px;font-weight:700;color:var(--primary);line-height:1.1;margin-bottom:4px}
.kpi .delta{font-size:11px;font-weight:600}
.kpi .delta.pos{color:var(--green)}.kpi .delta.neg{color:var(--red)}.kpi .delta.neu{color:var(--gray)}
.kpi .sub{font-size:11px;color:var(--gray);margin-top:4px}
.kpi .sub2{font-size:11px;color:var(--text);margin-top:3px}
.bar-bg{height:4px;background:var(--border);border-radius:4px;margin-top:10px;overflow:hidden}
.bar-fill{height:100%;border-radius:4px;transition:width .4s}
.kpi-tag{font-size:10px;font-weight:600;padding:2px 8px;border-radius:20px;display:inline-block;margin-top:4px}
.tag-aprov{background:rgba(7,106,118,.1);color:var(--primary)}
.tag-fin{background:rgba(65,168,185,.15);color:#076A76}
.tag-ren{background:rgba(251,174,75,.2);color:#a07030}
.kpi-divider{border:none;border-top:1px solid var(--border);margin:10px 0 8px}
.cac-btn{font-size:11px;text-decoration:underline;color:var(--primary);cursor:pointer;background:none;border:none;font-family:inherit;margin-top:2px;display:inline-block}
.modal-input{width:100%;border:1.5px solid var(--border);border-radius:8px;padding:7px 10px;font-size:13px;font-family:inherit;color:var(--text);background:#fff}
.row2{display:grid;grid-template-columns:2fr 1fr;gap:16px;margin-bottom:20px}
.row3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:20px}
.row-full{margin-bottom:20px}
.card{background:var(--card);border-radius:var(--r);padding:20px;box-shadow:var(--sh)}
.ctitle{font-size:12px;font-weight:700;letter-spacing:.8px;color:var(--text);margin-bottom:14px;
        display:flex;justify-content:space-between;align-items:center}
.ctitle select{font-size:11px;border:1px solid var(--border);border-radius:6px;padding:3px 8px;
               font-family:inherit;color:var(--text);cursor:pointer;background:#fff}
.pip-row{display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid var(--border)}
.pip-row:last-child{border-bottom:none}
.dot{width:9px;height:9px;border-radius:50%;flex-shrink:0;display:inline-block}
.pip-label{flex:1;font-size:12px}.pip-qtd{font-weight:700;font-size:13px;color:var(--primary)}
.pip-val{font-size:11px;color:var(--gray);margin-left:8px}
table{width:100%;border-collapse:collapse;font-size:12px}
thead tr{background:rgba(7,106,118,.06)}
th{padding:8px 10px;text-align:left;font-weight:700;font-size:10px;letter-spacing:.8px;color:var(--gray);
   text-transform:uppercase;border-bottom:2px solid var(--border)}
td{padding:7px 10px;border-bottom:1px solid var(--border)}
td.n{text-align:right;font-variant-numeric:tabular-nums}
tr:hover td{background:rgba(7,106,118,.03)}
.ano-row{margin-bottom:12px}
.ano-head{display:flex;align-items:baseline;gap:8px;margin-bottom:4px}
.ano-label{font-size:13px;font-weight:700;color:var(--text);min-width:38px}
.ano-val{font-size:14px;font-weight:700;color:var(--primary);flex:1}
.mix-notice{text-align:center;padding:30px;color:var(--gray);font-size:13px}
.modal-backdrop{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:200;align-items:center;justify-content:center}
.modal-backdrop.open{display:flex}
.modal{background:#fff;border-radius:12px;padding:28px;width:420px;max-width:95vw;max-height:80vh;overflow-y:auto;box-shadow:0 8px 40px rgba(0,0,0,.2)}
.modal h3{font-size:15px;font-weight:700;color:var(--primary);margin-bottom:6px}
.modal .desc{font-size:12px;color:var(--gray);margin-bottom:18px}
.cac-row{display:flex;align-items:center;gap:10px;margin-bottom:10px}
.cac-row label{flex:1;font-size:12px;font-weight:600;color:var(--text)}
.cac-row input{width:130px;border:1.5px solid var(--border);border-radius:6px;padding:5px 10px;font-size:12px;font-family:inherit}
.cac-row input:focus{outline:none;border-color:var(--primary)}
.btn-primary{background:var(--primary);color:#fff;border:none;border-radius:8px;padding:9px 20px;font-size:13px;font-weight:600;font-family:inherit;cursor:pointer;margin-top:12px}
.btn-primary:hover{background:#05565f}
.btn-sec{background:transparent;color:var(--gray);border:1.5px solid var(--border);border-radius:8px;padding:9px 16px;font-size:13px;font-family:inherit;cursor:pointer;margin-top:12px;margin-left:8px}
.footer{text-align:center;font-size:11px;color:var(--gray);padding:20px 0 8px;margin-top:10px}
@media(max-width:900px){.row2,.row3{grid-template-columns:1fr}.filter-bar{padding:10px 16px}.wrap{padding:16px 12px 40px}.hdr{padding:0 16px}}
"""

    js = f"""
const DATA = {data_json};
const COR_STATUS = {cor_status_js};
const COR_VERT   = {cor_vert_js};
const VERTS      = {verts_js};
const MESES_PT=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
const PIP_ORDER  = {pip_order_js};

const state = {{ year:'all', month:null, cac:{{}}, custom_period:false, from_ym:null, to_ym:null }};

const BRL  = v => 'R$ '+Number(v).toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
const BRLk = v => v>=1e6?'R$ '+(v/1e6).toFixed(1)+'M':v>=1e3?'R$ '+Math.round(v/1e3)+'k':BRL(v);
const dTag = (v,s='') => v==null?'<span class="delta neu">—</span>':v>=0?`<span class="delta pos">▲ ${{v.toFixed(1)}}%${{s}}</span>`:`<span class="delta neg">▼ ${{Math.abs(v).toFixed(1)}}%${{s}}</span>`;

const charts = {{}};
function killChart(id){{ if(charts[id]){{charts[id].destroy();delete charts[id];}} }}

function getFiltered(){{
  let s = DATA.serie;
  if(state.custom_period && state.from_ym && state.to_ym){{
    s=s.filter(d=>d.ym>=state.from_ym && d.ym<=state.to_ym);
  }} else {{
    if(state.year!=='all') s=s.filter(d=>d.ano===state.year);
    if(state.month!==null) s=s.filter(d=>d.mes===state.month);
  }}
  return s;
}}

function renderCards(){{
  const f=getFiltered(); if(!f.length) return;
  const last=f[f.length-1];
  const total=f.reduce((a,b)=>a+b.faturamento,0);
  const novos=f.reduce((a,b)=>a+b.novos,0);
  const recos=f.reduce((a,b)=>a+b.recorrentes,0);
  const isMulti=f.length>1;
  const pctMes=last.pct_meta||0;
  const mc=pctMes>=100?'var(--green)':pctMes>=70?'var(--accent)':'var(--red)';
  const mesesAcima=f.filter(d=>d.faturamento>=DATA.meta_mensal).length;
  const subFat=isMulti
    ? BRL(total)+'&nbsp; | &nbsp;'+mesesAcima+'/'+f.length+' meses &ge; break even'
    : pctMes.toFixed(1)+'% do break even ('+BRLk(DATA.meta_mensal)+'/mês)';
  document.getElementById('card-fat').innerHTML=`
    <label>${{isMulti?'Faturamento no Período':'Faturamento — '+last.label}}</label>
    <div class="v">${{BRLk(total)}}</div>
    ${{isMulti?'':dTag(last.mom,' vs mês ant.')}}
    <div class="bar-bg"><div class="bar-fill" style="width:${{Math.min(pctMes,100)}}%;background:${{mc}}"></div></div>
    <div class="sub">${{subFat}}</div>`;
  const anoRef=state.year!=='all'?state.year:new Date().getFullYear();
  const fav=parseFloat(DATA.fat_ano[anoRef]||0);
  const beTotal=getFiltered().reduce((s,d)=>s+getBreakeven(d.ym),0)||DATA.meta_mensal*12;
  const pctA=fav/beTotal*100;
  document.getElementById('card-acum').innerHTML=`
    <label>Acumulado ${{anoRef}}</label>
    <div class="v">${{BRLk(fav)}}</div>
    <div style="font-size:10px;color:var(--gray);margin:3px 0">Break Even: ${{BRL(DATA.meta_mensal)}}/m&ecirc;s &nbsp;&middot;&nbsp; Meta anual: ${{BRLk(DATA.meta_mensal*12)}}</div>
    <div class="bar-bg"><div class="bar-fill" style="width:${{Math.min(pctA,100)}}%;background:var(--accent)"></div></div>
    <div class="sub">${{pctA.toFixed(1)}}% da meta anual atingida</div>`;
  document.getElementById('card-clientes').innerHTML=`
    <label>Clientes no Período</label>
    <div class="v" style="font-size:20px;display:flex;gap:20px">
      <span><span style="font-size:11px;color:var(--gray);display:block">Novos</span>${{novos}}</span>
      <span><span style="font-size:11px;color:var(--gray);display:block">Antigos</span>${{recos}}</span>
    </div>
    <div class="sub" style="margin-top:8px">Total: ${{novos+recos}} atendimentos</div>`;
  // Pipeline agregado pelo período filtrado (não acumulado total)
  const pip={{}};
  f.forEach(s=>{{
    Object.entries(s.pipeline||{{}}).forEach(([st,qtd])=>{{
      if(!pip[st]) pip[st]={{qtd:0,valor:0}};
      pip[st].qtd+=qtd; pip[st].valor+=(s.pipeline_val||{{}})[st]||0;
    }});
  }});
  const sub={{}};
  f.forEach(s=>{{
    Object.entries(s.pip_sub||{{}}).forEach(([fase,v])=>{{
      if(!sub[fase]) sub[fase]={{qtd:0,valor:0}};
      sub[fase].qtd+=v.qtd; sub[fase].valor+=v.valor;
    }});
  }});
  const aprov=(pip['Aprovada']||{{}}).qtd||0;
  const aprov_v=(pip['Aprovada']||{{}}).valor||0;
  const qA=(sub['PROPOSTA APROVADA']||{{}}).qtd||0;
  const qF=(sub['PROPOSTA FINALIZADA']||{{}}).qtd||0;
  const qR=(sub['PROPOSTA RENOVADA (CASA)']||{{}}).qtd||0;
  // Taxa de conversão dinâmica (ERP + preliminares)
  const filterYmsC=new Set(f.map(s=>s.ym));
  let prelimAprov=0,prelimAprovV=0,prelimTot=0,prelimTotV=0;
  let prelimEmN=0,prelimEmNV=0,prelimEnv=0,prelimEnvV=0;
  loadPrelim();
  Object.entries(state.prelim||{{}}).forEach(([ym,arr])=>{{
    if(filterYmsC.size&&!filterYmsC.has(ym))return;
    arr.forEach(p=>{{
      const pv=parseFloat(p.valor||0);
      prelimTot++;prelimTotV+=pv;
      if(p.status==='Aprovada'){{prelimAprov++;prelimAprovV+=pv;}}
      if(p.status==='Em negociação'){{prelimEmN++;prelimEmNV+=pv;}}
      if(p.status==='Enviada'){{prelimEnv++;prelimEnvV+=pv;}}
    }});
  }});
  const erpTot=Object.values(pip).reduce((a,b)=>a+b.qtd,0);
  const erpTotV=Object.values(pip).reduce((a,b)=>a+b.valor,0);
  const denomQ=erpTot+prelimTot;const denomV=erpTotV+prelimTotV;
  const numQ=aprov+prelimAprov;const numV=aprov_v+prelimAprovV;
  const taxaQ=denomQ>0?(numQ/denomQ*100).toFixed(1):'—';
  const taxaV=denomV>0?(numV/denomV*100).toFixed(1):'—';
  document.getElementById('card-propostas').innerHTML=`
    <label>Propostas Aprovadas${{isMulti?' — '+f[0].ano:' — '+last.label}}</label>
    <div class="v">${{aprov+prelimAprov}}</div>
    <div style="margin:4px 0 3px">
      <span class="kpi-tag tag-aprov">Aprovada: ${{qA}}</span>
      <span class="kpi-tag tag-fin" style="margin-left:3px">Finalizada: ${{qF}}</span>
      <span class="kpi-tag tag-ren" style="margin-left:3px">Renovada: ${{qR}}</span>
    </div>
    <hr class="kpi-divider">
    <div class="sub2" style="display:flex;justify-content:space-between;margin-bottom:3px">
      <span style="color:var(--gray);font-size:11px">Taxa conversão (qtd)</span>
      <strong style="color:var(--primary)">${{taxaQ}}%</strong>
    </div>
    <div class="sub2" style="display:flex;justify-content:space-between;margin-bottom:2px">
      <span style="color:var(--gray);font-size:11px">Taxa conversão (valor)</span>
      <strong style="color:var(--primary)">${{taxaV}}%</strong>
    </div>
    <div class="sub2" style="display:flex;justify-content:space-between;font-size:10px;color:var(--gray)">
      <span>${{BRLk(numV)}} aprovado</span>
      <span>de ${{BRLk(denomV)}} cotado</span>
    </div>`;
  const emN=(pip['Em negociação']||{{}}).qtd||0;
  const emNV=(pip['Em negociação']||{{}}).valor||0;
  const env=(pip['Enviada']||{{}}).qtd||0;
  const envV=(pip['Enviada']||{{}}).valor||0;
  const totAbertas=emN+env+prelimEmN+prelimEnv;
  const totAbertasV=emNV+envV+prelimEmNV+prelimEnvV;
  document.getElementById('card-negociacao').innerHTML=`
    <label>Propostas Abertas${{isMulti?' — '+f[0].ano:' — '+last.label}}</label>
    <div class="v" style="color:var(--accent)">${{totAbertas}}</div>
    <div style="font-size:10px;color:var(--gray);margin:4px 0 2px;font-weight:700;text-transform:uppercase">ERP</div>
    <div class="sub2">Em negociação: <strong>${{emN}}</strong> &nbsp; Enviadas: <strong>${{env}}</strong></div>
    ${{(prelimEmN+prelimEnv)>0?`<div style=\"font-size:10px;color:var(--gray);margin:4px 0 2px;font-weight:700;text-transform:uppercase\">Fora do ERP</div><div class=\"sub2\">Em negociação: <strong>${{prelimEmN}}</strong> &nbsp; Enviadas: <strong>${{prelimEnv}}</strong></div>`:'' }}
    <hr class="kpi-divider">
    <div class="sub" style="margin-top:3px">Valor total: ${{BRLk(totAbertasV)}}</div>`;
  renderCacCard(f);
}}

function loadCacStorage(){{
  try{{state.cac=JSON.parse(localStorage.getItem('bb_cac')||'{{}}');}}catch(e){{state.cac={{}};}}
}}
function saveCac(){{
  document.querySelectorAll('.cac-input').forEach(inp=>{{
    const ym=inp.dataset.ym,val=parseFloat(inp.value.replace(',','.'));
    if(!isNaN(val)&&val>0) state.cac[ym]=val; else delete state.cac[ym];
  }});
  try{{localStorage.setItem('bb_cac',JSON.stringify(state.cac));}}catch(e){{}}
  closeCacModal(); renderCacCard(getFiltered());
}}
function openCacModal(){{
  const f=getFiltered();
  const rows=f.slice(-12).map(s=>`<div class="cac-row">
    <label>${{s.label}}</label>
    <input class="cac-input" type="number" data-ym="${{s.ym}}" value="${{state.cac[s.ym]||''}}" placeholder="0.00">
  </div>`).join('');
  document.getElementById('cac-inputs').innerHTML=rows||'<p>Nenhum mês no período.</p>';
  document.getElementById('cac-modal').classList.add('open');
}}
function closeCacModal(){{document.getElementById('cac-modal').classList.remove('open');}}

// ── Break-even por mês ───────────────────────────────────────────────────────
function loadBreakevenStorage(){{try{{state.breakeven=JSON.parse(localStorage.getItem('bb_breakeven')||'{{}}');}}catch(e){{state.breakeven={{}};}}}}
function saveBreakeven(){{
  document.querySelectorAll('.be-input').forEach(inp=>{{
    const ym=inp.dataset.ym,val=parseFloat(inp.value.replace(',','.'));
    if(!isNaN(val)&&val>0) state.breakeven[ym]=val; else delete state.breakeven[ym];
  }});
  try{{localStorage.setItem('bb_breakeven',JSON.stringify(state.breakeven));}}catch(e){{}}
  closeBreakevenModal(); renderFatChart(); renderFatFin();
}}
function openBreakevenModal(){{
  loadBreakevenStorage();
  const f=getFiltered();
  const rows=f.slice(-12).map(s=>`<div class="cac-row">
    <label>${{s.label}}</label>
    <input class="be-input" type="number" data-ym="${{s.ym}}" value="${{(state.breakeven&&state.breakeven[s.ym])||''}}" placeholder="${{DATA.meta_mensal}}">
  </div>`).join('');
  document.getElementById('be-inputs').innerHTML=rows||'<p>Nenhum mês no período.</p>';
  document.getElementById('breakeven-modal').classList.add('open');
}}
function closeBreakevenModal(){{document.getElementById('breakeven-modal').classList.remove('open');}}
function getBreakeven(ym){{return(state.breakeven&&state.breakeven[ym])||DATA.meta_mensal;}}
function renderCacCard(f){{
  let inv=0,novos=0;
  f.forEach(s=>{{inv+=parseFloat(state.cac[s.ym]||0);novos+=s.novos;}});
  const cac=inv>0&&novos>0?inv/novos:0;
  document.getElementById('cac-value').textContent=BRL(cac);
  document.getElementById('cac-sub').textContent=inv>0?`Invest.: ${{BRL(inv)}} · ${{novos}} novos clientes`:'Clique abaixo para informar';
}}

function renderFatChart(){{
  const f=getFiltered();
  killChart('fat');
  charts.fat=new Chart(document.getElementById('cFat').getContext('2d'),{{
    data:{{labels:f.map(d=>d.label),datasets:[
      {{type:'bar',label:'Faturamento',data:f.map(d=>d.faturamento),
        backgroundColor:f.map(d=>d.faturamento>=getBreakeven(d.ym)?'rgba(7,106,118,.75)':'rgba(251,174,75,.75)'),borderRadius:4,order:2}},
      {{type:'line',label:'Break Even',data:f.map(d=>getBreakeven(d.ym)),
        borderColor:'#DC3545',borderWidth:1.5,borderDash:[5,3],pointRadius:3,fill:false,order:1}},
      {{type:'line',label:'Meta',data:f.map(d=>(state.metas&&state.metas[d.ym])||null),
        borderColor:'#076A76',borderWidth:2,borderDash:[4,2],pointRadius:4,fill:false,order:0,spanGaps:false}},
    ]}},
    options:{{responsive:true,maintainAspectRatio:true,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>BRL(c.parsed.y)}}}}}},
      scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},
               y:{{ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
  }});
}}

function renderPipChart(){{
  // Agrega pipeline pelo período filtrado
  const f=getFiltered();
  const pip={{}};
  f.forEach(s=>{{
    Object.entries(s.pipeline||{{}}).forEach(([st,qtd])=>{{
      if(!pip[st]) pip[st]={{qtd:0,valor:0}};
      pip[st].qtd+=qtd; pip[st].valor+=(s.pipeline_val||{{}})[st]||0;
    }});
  }});
  const labels=PIP_ORDER.filter(k=>pip[k]&&pip[k].qtd>0);
  const vals=labels.map(k=>pip[k].qtd);
  const cors=labels.map(k=>COR_STATUS[k]||'#999');
  killChart('pip');
  if(!labels.length) return;
  charts.pip=new Chart(document.getElementById('cPip').getContext('2d'),{{
    type:'doughnut',data:{{labels,datasets:[{{data:vals,backgroundColor:cors,borderWidth:2,borderColor:'#fff'}}]}},
    options:{{responsive:true,cutout:'62%',plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>`${{c.label}}: ${{c.parsed}}`}}}}}}}}
  }});
  document.getElementById('pip-legend').innerHTML=labels.map((l,i)=>{{
    const v=pip[l];return`<div class="pip-row">
      <span class="dot" style="background:${{cors[i]}}"></span>
      <span class="pip-label">${{l}}</span><span class="pip-qtd">${{vals[i]}}</span>
      <span class="pip-val">${{BRLk(v.valor)}}</span></div>`;
  }}).join('');
}}

function renderVertChart(){{
  const f=getFiltered();
  const n=parseInt(document.getElementById('sel-n-meses').value)||0;
  const s=n>0?f.slice(-n):f;
  killChart('vert');
  const totalDs={{type:'line',label:'Total',data:s.map(d=>d.faturamento),
    borderColor:'#2c2c2c',borderWidth:2,pointRadius:3,pointBackgroundColor:'#2c2c2c',
    fill:false,order:0,yAxisID:'y'}};
  const barDs=VERTS.map(v=>({{label:v,data:s.map(d=>d.por_vertical[v]||0),
    backgroundColor:COR_VERT[v],borderWidth:0,borderRadius:2,type:'bar',order:1}}));
  charts.vert=new Chart(document.getElementById('cVert').getContext('2d'),{{
    data:{{labels:s.map(d=>d.label),datasets:[totalDs,...barDs]}},
    options:{{responsive:true,
      plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}},boxWidth:12,padding:12,
        filter:item=>item.text!=='Total'}}}},tooltip:{{callbacks:{{label:c=>` ${{c.dataset.label}}: ${{BRL(c.parsed.y)}}`}}}}}},
      scales:{{x:{{stacked:true,ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},y:{{stacked:true,ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
  }});
}}


function renderInadimplencia(){{
  const el=document.getElementById('inadim-section');
  if(!el)return;
  // Merge DATA (gerado pelo script) com localStorage (adicionados pelo user na sessão)
  const stored=JSON.parse(localStorage.getItem('inadim_v2')||'[]');
  const base=(DATA.inadimplencia_v2||[]);
  // Merge por id: localStorage prevalece para updates, itens novos somam
  const merged=[...base];
  stored.forEach(s=>{{
    if(!merged.find(b=>b.id===s.id)) merged.push(s);
  }});
  const f=getFiltered();
  const filterYms=new Set(f.map(d=>d.ym));
  const rows=merged.filter(r=>!filterYms.size||filterYms.has(r.ym))
    .sort((a,b)=>a.ym>b.ym?1:a.ym<b.ym?-1:0);
  // Totais por mês
  const totMes={{}};
  rows.forEach(r=>{{
    if(!totMes[r.ym])totMes[r.ym]=0;
    totMes[r.ym]+=parseFloat(r.valor||0);
  }});
  const totalGeral=rows.reduce((a,r)=>a+parseFloat(r.valor||0),0);
  const COR_VERT={{'Eventos':'#076A76','Casa':'#41A8B9','Escritório':'#2E86AB','Concierge':'#FBAE4B','Seminovos':'#B0865A'}};
  const MESES_ABR=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
  const fmtYm=ym=>{{const[y,m]=ym.split('-');return MESES_ABR[parseInt(m)-1]+'/'+y.slice(2);}};

  if(!rows.length){{
    el.innerHTML=`<div class="card"><div style="color:var(--gray);font-size:12px;padding:8px">Nenhuma inadimplência registrada no período.</div></div>`;
    return;
  }}
  // Tabela
  let html=`<div class="card"><div style="overflow-x:auto">`;
  html+=`<table style="width:100%;border-collapse:collapse;font-size:12px">`;
  html+=`<thead><tr style="background:#f8f9fa;border-bottom:2px solid var(--border)">
    <th style="padding:8px 10px;text-align:left;font-weight:700;color:var(--text)">Mês</th>
    <th style="padding:8px 10px;text-align:left;font-weight:700;color:var(--text)">Razão Social</th>
    <th style="padding:8px 10px;text-align:left;font-weight:700;color:var(--text)">Vertical</th>
    <th style="padding:8px 10px;text-align:left;font-weight:700;color:var(--text)">Contrato</th>
    <th style="padding:8px 10px;text-align:right;font-weight:700;color:var(--text)">Valor</th>
    <th style="padding:8px 10px;text-align:left;font-weight:700;color:var(--text)">Observações</th>
    <th style="padding:8px 10px;text-align:center;font-weight:700;color:var(--text)">Ação</th>
  </tr></thead><tbody>`;
  let lastYm='';
  rows.forEach(r=>{{
    const isNewMes=r.ym!==lastYm;
    if(isNewMes&&lastYm){{
      html+=`<tr style="background:#fff8ee;border-top:1px solid #FBAE4B20">
        <td colspan="4" style="padding:6px 10px;font-size:11px;font-weight:700;color:#FBAE4B;text-align:right">Subtotal ${{fmtYm(lastYm)}}</td>
        <td style="padding:6px 10px;text-align:right;font-weight:800;color:#DC3545">${{BRL(totMes[lastYm]||0)}}</td>
        <td colspan="2"></td></tr>`;
    }}
    lastYm=r.ym;
    const vCor=COR_VERT[r.vertical]||'#999';
    html+=`<tr style="border-bottom:1px solid var(--border)" data-id="${{r.id}}">
      <td style="padding:7px 10px">${{fmtYm(r.ym)}}</td>
      <td style="padding:7px 10px;font-weight:600">${{r.razao_social||''}}</td>
      <td style="padding:7px 10px"><span style="background:${{vCor}}22;color:${{vCor}};border-radius:4px;padding:2px 7px;font-size:10px;font-weight:700">${{r.vertical||''}}</span></td>
      <td style="padding:7px 10px;color:var(--gray)">${{r.contrato||'—'}}</td>
      <td style="padding:7px 10px;text-align:right;font-weight:700;color:#DC3545">${{BRL(parseFloat(r.valor||0))}}</td>
      <td style="padding:7px 10px;color:var(--gray);font-size:11px">${{r.obs||''}}</td>
      <td style="padding:7px 10px;text-align:center"><button onclick="deleteInadim('${{r.id}}')" style="background:none;border:none;cursor:pointer;color:#DC3545;font-size:13px" title="Remover">✕</button></td>
    </tr>`;
  }});
  // Subtotal do último mês
  if(lastYm){{
    html+=`<tr style="background:#fff8ee;border-top:1px solid #FBAE4B20">
      <td colspan="4" style="padding:6px 10px;font-size:11px;font-weight:700;color:#FBAE4B;text-align:right">Subtotal ${{fmtYm(lastYm)}}</td>
      <td style="padding:6px 10px;text-align:right;font-weight:800;color:#DC3545">${{BRL(totMes[lastYm]||0)}}</td>
      <td colspan="2"></td></tr>`;
  }}
  // Total Geral
  html+=`<tr style="background:#DC354510;border-top:2px solid #DC3545">
    <td colspan="4" style="padding:8px 10px;font-weight:800;color:#DC3545;text-align:right">Total Geral de Inadimplência</td>
    <td style="padding:8px 10px;text-align:right;font-weight:900;color:#DC3545;font-size:14px">${{BRL(totalGeral)}}</td>
    <td colspan="2"></td></tr>`;
  html+=`</tbody></table></div></div>`;
  el.innerHTML=html;
}}

function deleteInadim(id){{
  if(!confirm('Remover este registro de inadimplência?'))return;
  // Remove do localStorage
  const stored=JSON.parse(localStorage.getItem('inadim_v2')||'[]');
  const updated=stored.filter(r=>r.id!==id);
  localStorage.setItem('inadim_v2',JSON.stringify(updated));
  // Marca como deletado para exportação
  const deleted=JSON.parse(localStorage.getItem('inadim_deleted')||'[]');
  if(!deleted.includes(id))deleted.push(id);
  localStorage.setItem('inadim_deleted',JSON.stringify(deleted));
  renderInadimplencia();
}}

function openInadimModal(){{
  document.getElementById('inadim-modal').style.display='flex';
  const today=new Date();
  document.getElementById('inadim-ym').value=today.toISOString().slice(0,7);
}}
function closeInadimModal(){{document.getElementById('inadim-modal').style.display='none';}}
function saveInadim(){{
  const ym=document.getElementById('inadim-ym').value;
  const razao=document.getElementById('inadim-razao').value.trim();
  const vertical=document.getElementById('inadim-vertical').value;
  const contrato=document.getElementById('inadim-contrato').value.trim();
  const valor=parseFloat(document.getElementById('inadim-valor').value.replace(',','.'));
  const obs=document.getElementById('inadim-obs').value.trim();
  if(!ym||!razao||!vertical||isNaN(valor)||valor<=0){{alert('Preencha os campos obrigatórios (mês, razão social, vertical e valor).');return;}}
  const entry={{id:'inadim_'+Date.now(),ym,razao_social:razao,vertical,contrato,valor,obs,data_lancamento:new Date().toISOString().slice(0,10)}};
  const stored=JSON.parse(localStorage.getItem('inadim_v2')||'[]');
  stored.push(entry);
  localStorage.setItem('inadim_v2',JSON.stringify(stored));
  // Limpar form
  ['inadim-razao','inadim-contrato','inadim-valor','inadim-obs'].forEach(id=>{{document.getElementById(id).value='';}});
  closeInadimModal();
  renderInadimplencia();
}}
function exportManualInput(){{
  // Gera manual_input.json atualizado com inadimplência do localStorage
  const base=DATA.inadimplencia_v2||[];
  const stored=JSON.parse(localStorage.getItem('inadim_v2')||'[]');
  const deleted=JSON.parse(localStorage.getItem('inadim_deleted')||'[]');
  const merged=[...base.filter(r=>!deleted.includes(r.id))];
  stored.forEach(s=>{{if(!merged.find(b=>b.id===s.id))merged.push(s);}});
  const mi={{marketing:DATA.serie?Object.fromEntries(DATA.serie.filter(s=>s.inv_mkt_hist>0).map(s=>[s.ym,s.inv_mkt_hist])):{{}},inadimplencia:{{}},inadimplencia_v2:merged,propostas_preliminares:{{}},faturamento_manual:{{}}}};
  const blob=new Blob([JSON.stringify(mi,null,2)],{{type:'application/json'}});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='manual_input.json';a.click();
}}

function renderMix(){{
  const f=getFiltered();const mix=DATA.mix_por_mes;
  const hasMix=f.some(s=>mix[s.ym]&&(mix[s.ym].locacao+mix[s.ym].frete)>0);
  const ct=document.getElementById('mix-content');
  if(!hasMix){{killChart('mix');ct.innerHTML='<div class="mix-notice">Dados de mix disponíveis a partir de 2026 (Eloca).</div>';return;}}
  const m=f.filter(s=>mix[s.ym]||(s.por_vertical['Concierge']||0)>0||(s.por_vertical['Seminovos']||0)>0);
  ct.innerHTML='<canvas id="cMix" height="55"></canvas>';
  killChart('mix');
  charts.mix=new Chart(document.getElementById('cMix').getContext('2d'),{{
    type:'bar',
    data:{{labels:m.map(s=>s.label),datasets:[
      {{label:'Locação',data:m.map(s=>(mix[s.ym]||{{}}).locacao||0),backgroundColor:'rgba(7,106,118,.75)',borderRadius:4}},
      {{label:'Serviço/Frete',data:m.map(s=>(mix[s.ym]||{{}}).frete||0),backgroundColor:'rgba(251,174,75,.75)',borderRadius:4}},
      {{label:'Concierge',data:m.map(s=>s.por_vertical['Concierge']||0),backgroundColor:'rgba(251,174,75,.45)',borderRadius:4}},
      {{label:'Seminovos',data:m.map(s=>s.por_vertical['Seminovos']||0),backgroundColor:'rgba(176,134,90,.65)',borderRadius:4}},
    ]}},
    options:{{responsive:true,
      plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}},boxWidth:12}}}},tooltip:{{callbacks:{{label:c=>` ${{c.dataset.label}}: ${{BRL(c.parsed.y)}}`}}}}}},
      scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},y:{{ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
  }});
}}

function renderTicket(){{
  const mode=document.getElementById('sel-ticket').value;const f=getFiltered();
  let data;
  if(mode==='all'){{data=DATA.ticket;}}
  else if(mode==='month'&&f.length){{data=DATA.ticket_por_mes[f[f.length-1].ym]||DATA.ticket;}}
  else{{
    const agg={{}};
    VERTS.forEach(v=>{{let tot=0,cnt=0;f.forEach(s=>{{const t=DATA.ticket_por_mes[s.ym];if(t&&t[v]){{tot+=t[v].total;cnt+=t[v].contratos;}}}});agg[v]={{total:tot,contratos:cnt,ticket_medio:cnt?tot/cnt:0}};}});
    data=agg;
  }}
  document.getElementById('ticket-tbody').innerHTML=VERTS.map(v=>{{const t=data[v]||{{}};return`<tr>
    <td><span class="dot" style="background:${{COR_VERT[v]}}"></span> ${{v}}</td>
    <td class="n">${{t.contratos||0}}</td><td class="n">${{BRL(t.ticket_medio||0)}}</td><td class="n">${{BRL(t.total||0)}}</td>
  </tr>`;}}).join('');
}}

function renderTop(){{
  const mode=document.getElementById('sel-top').value;const f=getFiltered();
  let cl;
  if(mode==='all'){{cl=DATA.top_clientes.slice(0,10);}}
  else if(mode==='month'&&f.length){{cl=(DATA.top_por_mes[f[f.length-1].ym]||[]).slice(0,10);}}
  else{{
    const agg={{}};
    f.forEach(s=>{{(DATA.top_por_mes[s.ym]||[]).forEach(c=>{{if(!agg[c.nome])agg[c.nome]={{valor:0,contratos:0}};agg[c.nome].valor+=c.valor;agg[c.nome].contratos+=c.contratos;}});}});
    cl=Object.entries(agg).map(([n,v])=>({{nome:n,valor:v.valor,contratos:v.contratos}})).sort((a,b)=>b.valor-a.valor).slice(0,10);
  }}
  killChart('top');if(!cl.length)return;
  charts.top=new Chart(document.getElementById('cTop').getContext('2d'),{{
    type:'bar',
    data:{{labels:cl.map(c=>c.nome.length>28?c.nome.slice(0,28)+'…':c.nome),datasets:[{{data:cl.map(c=>c.valor),backgroundColor:'rgba(7,106,118,.7)',borderRadius:4}}]}},
    options:{{indexAxis:'y',responsive:true,maintainAspectRatio:true,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>BRL(c.parsed.x)}}}}}},
      scales:{{x:{{ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}},y:{{ticks:{{font:{{size:10}}}}}}}}}}
  }});
}}

function renderAnos(){{
  const anos=Object.keys(DATA.fat_ano);const maxV=Math.max(...Object.values(DATA.fat_ano));
  document.getElementById('fat-anos').innerHTML=anos.map((ano,i)=>{{
    const v=DATA.fat_ano[ano];const pct=maxV?(v/maxV*100).toFixed(1):0;
    const prev=i>0?DATA.fat_ano[anos[i-1]]:null;
    const yoy=prev?((v-prev)/prev*100).toFixed(1):null;
    const d=yoy==null?'<span class="delta neu">—</span>':yoy>=0?`<span class="delta pos">▲ ${{yoy}}%</span>`:`<span class="delta neg">▼ ${{Math.abs(yoy)}}%</span>`;
    return`<div class="ano-row"><div class="ano-head"><span class="ano-label">${{ano}}</span><span class="ano-val">${{BRL(v)}}</span>${{d}}</div>
    <div class="bar-bg"><div class="bar-fill" style="width:${{pct}}%;background:var(--primary)"></div></div></div>`;
  }}).join('');
}}

function renderFreq(){{
  const sortBy=state.freqSort||'locacoes';
  // Agrega freq_por_mes para os YMs do período filtrado
  const f=getFiltered();
  const agg={{}};
  f.forEach(s=>{{
    const mes=DATA.freq_por_mes[s.ym]||[];
    mes.forEach(c=>{{
      if(!agg[c.nome])agg[c.nome]={{nome:c.nome,locacoes:0,valor:0,vertical:c.vertical}};
      agg[c.nome].locacoes+=c.locacoes;
      agg[c.nome].valor+=c.valor;
    }});
  }});
  const sorted=Object.values(agg).sort((a,b)=>b[sortBy]-a[sortBy]).slice(0,30);
  document.getElementById('freq-sort-btn').textContent=sortBy==='locacoes'?'Ordenar por Receita':'Ordenar por Locações';
  document.getElementById('freq-sort-col').textContent=sortBy==='locacoes'?'Locações':'Receita';
  document.getElementById('freq-tbody').innerHTML=sorted.map((c,i)=>{{
    const tm=c.locacoes?c.valor/c.locacoes:0;
    const destVal=sortBy==='locacoes'?`<td class="n" style="font-weight:700;color:var(--primary)">${{c.locacoes}}</td><td class="n">${{BRL(c.valor)}}</td>`
                                     :`<td class="n">${{c.locacoes}}</td><td class="n" style="font-weight:700;color:var(--primary)">${{BRL(c.valor)}}</td>`;
    return`<tr><td style="color:var(--gray);font-weight:700;text-align:center">${{i+1}}</td>`
      +`<td>${{c.nome||'—'}}</td>`
      +`<td style="font-size:10px;color:var(--gray)">${{c.vertical||'—'}}</td>`
      +destVal
      +`<td class="n">${{BRL(tm)}}</td></tr>`;
  }}).join('');
}}
function toggleFreqSort(){{
  state.freqSort=(state.freqSort||'locacoes')==='locacoes'?'valor':'locacoes';
  renderFreq();
}}

function updateDesc(){{
  const MN=['','Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'];
  let desc,kpilbl;
  if(state.custom_period&&state.from_ym&&state.to_ym){{
    desc=`${{state.from_ym}} a ${{state.to_ym}}`;kpilbl=`Performance Base ERP — ${{state.from_ym}} a ${{state.to_ym}}`;
  }} else {{
    desc=state.month?`${{MN[state.month]}}/${{state.year}}`:state.year!=='all'?`Ano ${{state.year}}`:'Histórico completo';
    kpilbl=state.month?`Performance Base ERP — ${{MN[state.month]}}/${{state.year}}`:state.year!=='all'?`Performance Base ERP — ${{state.year}}`:'Performance Base ERP — Histórico';
  }}
  document.getElementById('filter-desc').textContent=`Exibindo: ${{desc}}`;
  document.getElementById('kpi-sec-label').textContent=kpilbl;
}}


// ── Evolução de Clientes — state e helpers ───────────────────────────────────
const cliState = {{
  erp:{{year:'all',month:null,filtro:'ambos'}},
  crm:{{year:'all',month:null,filtro:'ambos'}},
  man:{{year:'all',month:null,filtro:'ambos'}}
}};
const VERT_ABBR = {{'EVENTOS':'Eventos','LOCAÇÃO EVENTOS':'Eventos','CASA':'Casa','LOCAÇÃO CASA':'Casa',
  'ESCRITÓRIO':'Escritório','LOCAÇÃO ESCRITÓRIO':'Escritório','CONCIERGE':'Concierge',
  'SEMINOVOS':'Seminovos','VENDA DE SEMINOVOS':'Seminovos','OUTRAS RECEITAS':'Outras','COMISSÃO':'Outras'}};
function normVert(v){{return VERT_ABBR[(v||'').toUpperCase()]||(v||'');}}

function clientCard(nome, vertical, tipo){{
  const COR_N='#076A76', COR_A='#FBAE4B';
  const cor = tipo==='Novo' ? COR_N : COR_A;
  return `<div style="background:${{cor}}18;border:1.5px solid ${{cor}}66;border-radius:8px;padding:8px 12px;min-width:110px;max-width:180px;box-sizing:border-box">
    <div style="font-size:9px;font-weight:800;color:${{cor}};text-transform:uppercase;letter-spacing:.8px">${{tipo}}</div>
    <div style="font-size:11px;font-weight:600;color:#222;line-height:1.3;margin-top:2px">${{nome}}</div>
    ${{vertical?`<div style="font-size:10px;color:var(--gray);margin-top:2px">${{normVert(vertical)}}</div>`:''}}
  </div>`;
}}

function buildCardsHTML(det, filtro){{
  const novos=det.novos||[], antigos=det.antigos||[];
  const showN=(filtro==='novos'||filtro==='ambos'), showA=(filtro==='antigos'||filtro==='ambos');
  const items=[...(showN?novos:[]).map(c=>Object.assign({{tipo:'Novo'}},typeof c==='string'?{{nome:c,vertical:''}}:c)),
               ...(showA?antigos:[]).map(c=>Object.assign({{tipo:'Antigo'}},typeof c==='string'?{{nome:c,vertical:''}}:c))];
  if(!items.length) return `<p style="font-size:12px;color:var(--gray)">Nenhum cliente${{filtro==='ambos'?'':' '+filtro}} no período selecionado.</p>`;
  return `<div style="display:flex;flex-wrap:wrap;gap:8px">`+items.map(c=>clientCard(c.nome||c,c.vertical||'',c.tipo)).join('')+`</div>`;
}}

function _cliAgg(det, allYms, curYear, curMonth){{
  // Filtra yms pelo ano+mês selecionados e agrega clientes (sem duplicatas)
  const MN=['','Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
  let filtYms = allYms;
  if(curYear!=='all') filtYms=filtYms.filter(ym=>ym.startsWith(curYear+'-'));
  const mPad = curMonth ? String(curMonth).padStart(2,'0') : null;
  if(mPad) filtYms=filtYms.filter(ym=>ym.split('-')[1]===mPad);
  const agg={{novos:[],antigos:[]}};
  const sN=new Set(),sA=new Set();
  filtYms.forEach(ym=>{{
    const d=det[ym]||{{}};
    (d.novos||[]).forEach(c=>{{const k=(c.nome||c).toUpperCase();if(!sN.has(k)){{sN.add(k);agg.novos.push(c);}}}});
    (d.antigos||[]).forEach(c=>{{const k=(c.nome||c).toUpperCase();if(!sA.has(k)){{sA.add(k);agg.antigos.push(c);}}}});
  }});
  return agg;
}}

function mkCliControls(bKey, allYms, curYear, curMonth, curFiltro, novosCount, antigosCount){{
  const MN=['','Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
  // Anos disponíveis
  const years=[...new Set(allYms.map(ym=>ym.split('-')[0]))].sort();
  const yearBtns=[['all','Todos'],...years.map(y=>[y,y])].map(([y,lbl])=>{{
    const act=curYear===y;
    return `<button onclick="cliState['${{bKey}}'].year='${{y}}';cliState['${{bKey}}'].month=null;renderEvolucaoClientes()" style="font-size:10px;padding:3px 10px;border-radius:5px;border:1.5px solid ${{act?'#076A76':'#ddd'}};background:${{act?'#076A76':'#fff'}};color:${{act?'#fff':'#555'}};cursor:pointer;font-weight:${{act?700:400}}">${{lbl}}</button>`;
  }}).join('');
  // Meses disponíveis para o ano selecionado
  const filtYms = curYear==='all' ? allYms : allYms.filter(ym=>ym.startsWith(curYear+'-'));
  const monthOpts = filtYms.map(ym=>{{const[y,m]=ym.split('-');const mPad=m;const mVal=String(+m);return `<option value="${{mVal}}"${{curMonth===mVal?' selected':''}}>${{MN[+m]}}/${{y}}</option>`;}}).join('');
  const monthSel = `<select style="font-size:11px;padding:4px 8px;border:1px solid #ddd;border-radius:6px" onchange="cliState['${{bKey}}'].month=this.value||null;renderEvolucaoClientes()">
    <option value=""${{!curMonth?' selected':''}}>Todos os meses</option>${{monthOpts}}</select>`;
  // Filtro Novos/Antigos/Ambos
  const filterBtns=['novos','antigos','ambos'].map(f=>{{
    const act=curFiltro===f;
    const lbl=f==='novos'?'Novos':f==='antigos'?'Antigos':'Ambos';
    return `<button onclick="cliState['${{bKey}}'].filtro='${{f}}';renderEvolucaoClientes()" style="font-size:10px;padding:3px 10px;border-radius:5px;border:1.5px solid ${{act?'#076A76':'#ddd'}};background:${{act?'#076A76':'#fff'}};color:${{act?'#fff':'#555'}};cursor:pointer;font-weight:${{act?700:400}}">${{lbl}}</button>`;
  }}).join('');
  return `<div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;margin-bottom:14px;padding:10px 12px;background:#f8f9fa;border-radius:8px">
    <span style="font-size:10px;font-weight:700;color:var(--gray);text-transform:uppercase;letter-spacing:1px">Ano:</span>
    ${{yearBtns}}
    ${{monthSel}}
    <span style="font-size:10px;color:var(--border);margin:0 2px">|</span>
    ${{filterBtns}}
    ${{(()=>{{const tot=novosCount+antigosCount;const pN=tot?Math.round(novosCount/tot*100):0;const pA=tot?100-pN:0;return tot?`<span style="font-size:10px;color:var(--gray);margin-left:4px"><span style="color:#076A76;font-weight:700">${{novosCount}} novos (${{pN}}%)</span> · <span style="color:#FBAE4B;font-weight:700">${{antigosCount}} antigos (${{pA}}%)</span></span>`:'<span style="font-size:10px;color:var(--gray);margin-left:4px">Sem clientes no período</span>';}})()}}
  </div>`;
}}

function renderEvolucaoBase(bKey, det, elId){{
  const el=document.getElementById(elId); if(!el) return;
  const allYms=Object.keys(det).sort();
  if(!allYms.length){{el.innerHTML='<p style="font-size:12px;color:var(--gray)">Sem dados disponíveis.</p>';return;}}
  const curYear=cliState[bKey].year||'all';
  const curMonth=cliState[bKey].month||null;
  const curFiltro=cliState[bKey].filtro||'ambos';
  const agg=_cliAgg(det,allYms,curYear,curMonth);
  el.innerHTML=mkCliControls(bKey,allYms,curYear,curMonth,curFiltro,agg.novos.length,agg.antigos.length)+buildCardsHTML(agg,curFiltro);
}}

function renderEvolucaoManual(){{
  const el=document.getElementById('evol-cli-man'); if(!el) return;
  // allSeen: histórico completo ERP + CRM (para classificar novo/antigo)
  const allSeen=new Set();
  Object.values(DATA.erp_detalhes||{{}}).forEach(v=>{{
    (v.novos||[]).forEach(c=>allSeen.add((c.nome||c).trim().toUpperCase()));
    (v.antigos||[]).forEach(c=>allSeen.add((c.nome||c).trim().toUpperCase()));
  }});
  Object.values(DATA.crm_detalhes||{{}}).forEach(v=>{{
    (v.novos||[]).forEach(c=>allSeen.add((c.nome||c).trim().toUpperCase()));
    (v.antigos||[]).forEach(c=>allSeen.add((c.nome||c).trim().toUpperCase()));
  }});
  // Montar manDet a partir de state.prelim
  const manDet={{}};
  Object.entries(state.prelim||{{}}).forEach(([ym,arr])=>{{
    if(!manDet[ym]) manDet[ym]={{novos:[],antigos:[]}};
    arr.forEach(p=>{{
      const nome=(p.cliente||'').trim(), vert=p.vertical||''; if(!nome) return;
      const key=nome.toUpperCase();
      if(allSeen.has(key)) manDet[ym].antigos.push({{nome,vertical:vert}});
      else {{ manDet[ym].novos.push({{nome,vertical:vert}}); allSeen.add(key); }}
    }});
  }});
  const allYms=Object.keys(manDet).sort();
  if(!allYms.length){{el.innerHTML='<p style="font-size:12px;color:var(--gray)">Nenhum lançamento manual registrado.</p>';return;}}
  const curYear=cliState.man.year||'all';
  const curMonth=cliState.man.month||null;
  const curFiltro=cliState.man.filtro||'ambos';
  const agg=_cliAgg(manDet,allYms,curYear,curMonth);
  el.innerHTML=mkCliControls('man',allYms,curYear,curMonth,curFiltro,agg.novos.length,agg.antigos.length)+buildCardsHTML(agg,curFiltro);
}}

function renderEvolucaoClientes(){{
  renderEvolucaoBase('erp', DATA.erp_detalhes||{{}}, 'evol-cli-erp');
  renderEvolucaoBase('crm', DATA.crm_detalhes||{{}}, 'evol-cli-crm');
  renderEvolucaoManual();
}}

const STATUS_ORDER=['Aprovada','Enviada','Em negociação','Reprovada','Sem continuidade','Cancelada','Preliminar'];
const COR_ST={{'Aprovada':'#076A76','Enviada':'#41A8B9','Em negociação':'#FBAE4B',
  'Reprovada':'#DC3545','Sem continuidade':'#A9A69F','Cancelada':'#61605B','Preliminar':'#9B59B6'}};

function mkStatusGrid(pip,tot){{
  const rows=STATUS_ORDER.filter(k=>pip[k]&&pip[k].qtd>0);
  if(!rows.length) return '<p style="font-size:11px;color:var(--gray)">Nenhuma proposta.</p>';
  return '<div style="display:flex;flex-wrap:wrap;gap:8px">'
    +rows.map(k=>`<div style="background:#f8f9fa;border-radius:8px;padding:8px 12px;border-left:3px solid ${{COR_ST[k]}};min-width:100px">`
      +`<div style="font-size:9px;font-weight:700;color:var(--gray);text-transform:uppercase">${{k}}</div>`
      +`<div style="font-size:20px;font-weight:800;color:${{COR_ST[k]}}">${{pip[k].qtd}}</div>`
      +`<div style="font-size:10px;color:var(--gray)">${{BRLk(pip[k].valor)}}</div>`
    +'</div>').join('')+'</div>';
}}

function renderCrmBlock(){{
  const elBloco=document.getElementById('crm-bloco');
  const elLabel=document.getElementById('sec-crm-label');
  const crmMes=DATA.crm_pipeline_mes||{{}};
  const filterYms=new Set(getFiltered().map(d=>d.ym));
  const pip={{}};
  let totQ=0,totV=0;
  Object.entries(crmMes).forEach(([ym,statuses])=>{{
    if(filterYms.size&&!filterYms.has(ym))return;
    Object.entries(statuses).forEach(([st,v])=>{{
      if(!pip[st])pip[st]={{qtd:0,valor:0}};
      pip[st].qtd+=v.qtd; pip[st].valor+=v.valor;
      totQ+=v.qtd; totV+=v.valor;
    }});
  }});
  if(!totQ){{
    if(elBloco)elBloco.innerHTML='';
    if(elLabel)elLabel.style.display='none';
    return;
  }}
  if(elLabel)elLabel.style.display='';
  const periodoLabel=state.custom_period?`${{state.from_ym}} a ${{state.to_ym}}`:
    state.month?`${{['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'][state.month-1]}}/${{state.year}}`:
    state.year!=='all'?String(state.year):'Histórico';
  if(elBloco)elBloco.innerHTML=
    `<div class="card" style="margin-bottom:0">`
    +`<div style="font-size:11px;font-weight:700;color:#1a7abf;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;padding-bottom:4px;border-bottom:2px solid #1a7abf">`
    +`Performance Base CRM — ${{periodoLabel}}`
    +`<span style="float:right;font-weight:400;font-size:11px;color:var(--gray)">${{totQ}} propostas · ${{BRLk(totV)}}</span></div>`
    +mkStatusGrid(pip,totQ)
    +`<p style="font-size:10px;color:var(--gray);margin-top:8px">* Propostas APROVADA / FINALIZADA / RENOVADA CASA já contabilizadas na Base ERP</p>`
    +`</div>`;
}}
function renderCrmEntregas(){{
  const elEnt=document.getElementById('crm-entregas');
  const elLabel=document.getElementById('sec-crm-ent-label');
  const lst=DATA.crm_entregas_aberto||[];
  const filterYms=new Set(getFiltered().map(d=>d.ym));
  const rows=lst.filter(r=>!filterYms.size||filterYms.has(r.ym));
  if(!rows.length){{
    if(elEnt)elEnt.innerHTML='';
    if(elLabel)elLabel.style.display='none';
    return;
  }}
  if(elLabel)elLabel.style.display='';
  const fmtD=s=>s||'—';
  const COR_ST={{'Enviada':'#41A8B9','Em negociação':'#FBAE4B'}};
  const rows_html=rows.sort((a,b)=>a.ym>b.ym?1:-1).map(r=>
    `<tr>`
    +`<td style="font-size:11px">${{r.cliente}}</td>`
    +`<td style="font-size:11px">${{r.vertical}}</td>`
    +`<td style="font-size:11px"><span style="background:${{COR_ST[r.status]||'#ccc'}};color:#fff;border-radius:4px;padding:2px 6px;font-size:10px">${{r.status}}</span></td>`
    +`<td style="font-size:11px;text-align:center">${{fmtD(r.data_inicio)}}</td>`
    +`<td style="font-size:11px;text-align:center">${{fmtD(r.data_fim)}}</td>`
    +`<td style="font-size:11px;text-align:right;font-weight:700">${{BRL(r.valor)}}</td>`
    +`</tr>`
  ).join('');
  if(elEnt)elEnt.innerHTML=
    `<div class="card">`
    +`<div style="font-size:11px;font-weight:700;color:#1a7abf;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;padding-bottom:4px;border-bottom:2px solid #1a7abf">`
    +`Previsão de Entregas — Enviadas e Em Negociação (CRM)`
    +`<span style="float:right;font-weight:400;font-size:11px;color:var(--gray)">${{rows.length}} proposta(s)</span></div>`
    +`<p style="font-size:10px;color:var(--gray);margin-bottom:10px">Inclui apenas status Enviada e Em Negociação. Aprovadas/Finalizadas/Renovadas já estão na Base ERP.</p>`
    +`<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">`
    +`<thead><tr style="font-size:10px;color:var(--gray);text-transform:uppercase">`
    +`<th style="text-align:left;padding:4px 6px">Cliente</th>`
    +`<th style="text-align:left;padding:4px 6px">Vertical</th>`
    +`<th style="text-align:left;padding:4px 6px">Status</th>`
    +`<th style="text-align:center;padding:4px 6px">Início</th>`
    +`<th style="text-align:center;padding:4px 6px">Fim</th>`
    +`<th style="text-align:right;padding:4px 6px">Valor</th>`
    +`</tr></thead>`
    +`<tbody>${{rows_html}}</tbody>`
    +`</table></div>`
    +`</div>`;
}}
function renderAll(){{
  updateDesc();renderCards();renderFatChart();renderPipChart();renderFatManualList();renderCrmBlock();renderCrmEntregas();
  renderVertChart();renderMix();renderTicket();renderTop();renderAnos();renderFreq();
  renderTotalizador();renderMacro();renderEntregas();renderFatFin();renderPrelimList();renderInadimplencia();
  renderEvolucaoClientes();
}}

function loadMetasStorage(){{try{{state.metas=JSON.parse(localStorage.getItem('bb_metas')||'{{}}');}}catch(e){{state.metas={{}};}}}}
function openMetasModal(){{
  loadMetasStorage();
  const f=getFiltered();
  let html='';
  f.forEach(s=>{{
    const[y,m]=s.ym.split('-');
    const lbl=MESES_PT[parseInt(m)-1]+'/'+y.slice(2);
    const cur=(state.metas||{{}})[s.ym]||'';
    html+=`<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
      <span style="font-size:11px;color:var(--gray);width:62px">${{lbl}}</span>
      <input type="number" class="meta-input modal-input" data-ym="${{s.ym}}" value="${{cur}}" placeholder="150000" style="flex:1">
    </div>`;
  }});
  document.getElementById('metas-inputs').innerHTML=html||'<p style="color:var(--gray);font-size:12px">Selecione um período primeiro.</p>';
  document.getElementById('metas-modal').style.display='flex';
}}
function closeMetasModal(){{document.getElementById('metas-modal').style.display='none';}}
function saveMetas(){{
  if(!state.metas)state.metas={{}};
  document.querySelectorAll('.meta-input').forEach(inp=>{{
    const ym=inp.dataset.ym;const v=parseFloat(inp.value);
    if(!isNaN(v)&&v>0)state.metas[ym]=v; else delete state.metas[ym];
  }});
  try{{localStorage.setItem('bb_metas',JSON.stringify(state.metas));}}catch(e){{}}
  closeMetasModal();renderFatChart();
}}

function setFonteFilter(fonte,btn){{
  state.fonte_filter=fonte;
  document.querySelectorAll('.fonte-btn').forEach(b=>b.classList.remove('active'));
  if(btn)btn.classList.add('active');
  renderTotalizador();
}}

function renderTotalizador(){{
  const f=getFiltered();
  // ── ERP pipeline (apenas ERP, sem preliminares) ──
  const erpPip={{}};
  f.forEach(s=>{{
    Object.entries(s.pipeline||{{}}).forEach(([st,qtd])=>{{
      if(!erpPip[st])erpPip[st]={{qtd:0,valor:0}};
      erpPip[st].qtd+=qtd;
      erpPip[st].valor+=((s.pipeline_val||{{}})[st]||0);
    }});
  }});
  const erpTotQ=Object.values(erpPip).reduce((a,b)=>a+b.qtd,0);
  const erpTotV=Object.values(erpPip).reduce((a,b)=>a+b.valor,0);
  // ── Preliminares (fora do ERP) ──
  loadPrelim();
  const filterYms=new Set(f.map(s=>s.ym));
  const prelPip={{}};
  let prelTotQ=0,prelTotV=0;
  Object.entries(state.prelim||{{}}).forEach(([ym,arr])=>{{
    if(filterYms.size&&!filterYms.has(ym))return;
    arr.forEach(p=>{{
      const st=p.status||'Preliminar';
      if(!prelPip[st])prelPip[st]={{qtd:0,valor:0}};
      prelPip[st].qtd++;prelPip[st].valor+=parseFloat(p.valor||0);
      prelTotQ++;prelTotV+=parseFloat(p.valor||0);
    }});
  }});
  // ── CRM pipeline (para Total Geral) ──
  const crmMesData=DATA.crm_pipeline_mes||{{}};
  const crmPipAll={{}};
  Object.entries(crmMesData).forEach(([ym,statuses])=>{{
    if(filterYms.size&&!filterYms.has(ym))return;
    Object.entries(statuses).forEach(([st,v])=>{{
      if(!crmPipAll[st])crmPipAll[st]={{qtd:0,valor:0}};
      crmPipAll[st].qtd+=v.qtd; crmPipAll[st].valor+=v.valor;
    }});
  }});
  const fonteAtiva=state.fonte_filter||'todos';

  // ── Total consolidado real (sempre calculado para servir de base de %%) ──
  const consolidadoPip={{}};
  const _addConsolidado=(pip,soAprovERP)=>Object.entries(pip).forEach(([k,v])=>{{
    if(soAprovERP&&k==='Aprovada')return;
    if(!consolidadoPip[k])consolidadoPip[k]={{qtd:0,valor:0}};
    consolidadoPip[k].qtd+=v.qtd; consolidadoPip[k].valor+=v.valor;
  }});
  _addConsolidado(erpPip,false);
  _addConsolidado(crmPipAll,true);
  _addConsolidado(prelPip,true);
  const consolidadoTotQ=Object.values(consolidadoPip).reduce((a,b)=>a+b.qtd,0);
  const consolidadoTotV=Object.values(consolidadoPip).reduce((a,b)=>a+b.valor,0);

  // ── Pipeline da fonte ativa (para Total Geral adaptado) ──
  const allPip={{}};
  const _addPip=(pip,soAprovERP)=>Object.entries(pip).forEach(([k,v])=>{{
    if(soAprovERP&&k==='Aprovada')return;
    if(!allPip[k])allPip[k]={{qtd:0,valor:0}};
    allPip[k].qtd+=v.qtd; allPip[k].valor+=v.valor;
  }});
  if(fonteAtiva==='todos'||fonteAtiva==='erp')  _addPip(erpPip,false);
  if(fonteAtiva==='todos'||fonteAtiva==='crm')  _addPip(crmPipAll,true);
  if(fonteAtiva==='todos'||fonteAtiva==='manual')_addPip(prelPip,true);
  const allTotQ=Object.values(allPip).reduce((a,b)=>a+b.qtd,0);
  const allTotV=Object.values(allPip).reduce((a,b)=>a+b.valor,0);

  // ── Visibilidade dos blocos por origem (Opção A) ──
  const showErp    = fonteAtiva==='todos'||fonteAtiva==='erp';
  const showCrm    = fonteAtiva==='todos'||fonteAtiva==='crm';
  const showManual = fonteAtiva==='todos'||fonteAtiva==='manual';

  // Bloco ERP: usa o div prop-erp-bloco (já existe) + visibilidade
  const elErpWrap=document.getElementById('erp-bloco-wrap');
  if(elErpWrap) elErpWrap.style.display=showErp?'':'none';

  // Bloco CRM: usa sec-crm-label + crm-bloco
  const elCrmLabel=document.getElementById('sec-crm-label');
  const elCrmBloco=document.getElementById('crm-bloco');
  const elCrmEnt=document.getElementById('sec-crm-ent-label');
  const elCrmEntBloco=document.getElementById('crm-entregas');
  if(elCrmLabel) elCrmLabel.style.display=showCrm?'':'none';
  if(elCrmBloco) elCrmBloco.style.display=showCrm?'':'none';
  if(elCrmEnt) elCrmEnt.style.display=showCrm?'':'none';
  if(elCrmEntBloco) elCrmEntBloco.style.display=showCrm?'':'none';

  // Bloco Manual: sec-label Propostas Preliminares + prelim-list + prelim-totalizador
  document.querySelectorAll('.manual-bloco-wrap').forEach(el=>{{
    el.style.display=showManual?'':'none';
  }});

  // ── Card de contexto percentual quando fonte isolada ──
  const elCtx=document.getElementById('fonte-context-card');
  if(elCtx){{
    if(fonteAtiva==='todos'||!consolidadoTotQ){{
      elCtx.style.display='none';
    }} else {{
      const pctQ=consolidadoTotQ?Math.round(allTotQ/consolidadoTotQ*100):0;
      const pctV=consolidadoTotV?Math.round(allTotV/consolidadoTotV*100):0;
      const nomeFonte={{erp:'ERP',crm:'CRM',manual:'Lançamentos Manuais'}}[fonteAtiva]||fonteAtiva;
      const barQ=`<div style="background:#e8e8e8;border-radius:4px;height:8px;margin:4px 0 10px"><div style="background:var(--primary);width:${{pctQ}}%;height:8px;border-radius:4px;transition:.4s"></div></div>`;
      const barV=`<div style="background:#e8e8e8;border-radius:4px;height:8px;margin:4px 0 0"><div style="background:var(--accent);width:${{pctV}}%;height:8px;border-radius:4px;transition:.4s"></div></div>`;
      elCtx.style.display='';
      elCtx.innerHTML=
        `<div style="background:linear-gradient(135deg,#f0f9fa 0%,#e8f5f7 100%);border:1px solid var(--secondary);border-radius:10px;padding:14px 18px;margin-bottom:12px">`
        +`<div style="font-size:11px;font-weight:800;color:var(--primary);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px">`
        +`${{nomeFonte}} — Participação no Total Consolidado</div>`
        +`<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">`
        +`<div><div style="font-size:11px;color:var(--gray);margin-bottom:2px">Participação em Propostas</div>`
        +`<div style="font-size:26px;font-weight:900;color:var(--primary)">${{pctQ}}%</div>`
        +`<div style="font-size:10px;color:var(--gray)">${{allTotQ}} de ${{consolidadoTotQ}} propostas</div>`
        +barQ+`</div>`
        +`<div><div style="font-size:11px;color:var(--gray);margin-bottom:2px">Participação em Valor</div>`
        +`<div style="font-size:26px;font-weight:900;color:var(--accent)">${{pctV}}%</div>`
        +`<div style="font-size:10px;color:var(--gray)">${{BRLk(allTotV)}} de ${{BRLk(consolidadoTotV)}}</div>`
        +barV+`</div>`
        +`</div></div>`;
    }}
  }}

  // mkStatusGrid e COR_ST agora são globais
  // ── Renderiza Bloco ERP no topo (antes do lançamento manual) ──
  const elErp=document.getElementById('prop-erp-bloco');
  const periodoLabel=state.custom_period?`${{state.from_ym}} a ${{state.to_ym}}`:
    state.month?`${{['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'][state.month-1]}}/${{state.year}}`:
    state.year!=='all'?String(state.year):'Histórico';
  if(elErp) elErp.innerHTML=
    `<div class="card" style="margin-bottom:0">`
    +`<div style="font-size:11px;font-weight:700;color:var(--primary);text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;padding-bottom:4px;border-bottom:2px solid var(--primary)">`
    +`Performance Base ERP — ${{periodoLabel}}`
    +`<span style="float:right;font-weight:400;font-size:11px;color:var(--gray)">${{erpTotQ}} propostas · ${{BRLk(erpTotV)}}</span></div>`
    +mkStatusGrid(erpPip,erpTotQ)
    +`</div>`;
  // ── Renderiza Prelim + Total após lançamento manual ──
  const el=document.getElementById('prop-totalizador');
  if(!el)return;
  el.innerHTML=
    // Bloco Propostas Preliminares — fora do ERP
    `<div class="card" style="margin-bottom:12px">`
    +`<div style="font-size:11px;font-weight:700;color:#9B59B6;text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;padding-bottom:4px;border-bottom:2px solid #9B59B6">`
    +`Propostas Preliminares — Fora do ERP e do CRM`
    +`<span style="float:right;font-weight:400;font-size:11px;color:var(--gray)">${{prelTotQ}} propostas · ${{BRLk(prelTotV)}}</span></div>`
    +`${{prelTotQ>0?mkStatusGrid(prelPip,prelTotQ):'<p style="font-size:11px;color:var(--gray)">Nenhuma proposta preliminar no período.</p>'}}`
    +(()=>{{
      // Previsão de execução agrupada por mês de data_exec
      const execMap={{}};
      Object.entries(state.prelim||{{}}).forEach(([ym,arr])=>{{
        if(filterYms.size&&!filterYms.has(ym))return;
        arr.forEach(p=>{{
          if(!p.data_exec)return;
          const execYm=p.data_exec.slice(0,7); // YYYY-MM
          if(!execMap[execYm])execMap[execYm]={{qtd:0,valor:0}};
          execMap[execYm].qtd++;
          execMap[execYm].valor+=parseFloat(p.valor||0);
        }});
      }});
      const execEntries=Object.entries(execMap).sort((a,b)=>a[0]>b[0]?1:-1);
      if(!execEntries.length) return '';
      const MESES_ABR=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'];
      const fmtYm=ym=>{{const[y,m]=ym.split('-');return MESES_ABR[parseInt(m)-1]+'/'+y.slice(2);}};
      return `<div style="margin-top:14px;border-top:1px solid rgba(155,89,182,.25);padding-top:12px">`
        +`<div style="font-size:10px;font-weight:700;color:#9B59B6;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Previsão por Mês de Execução</div>`
        +`<div style="display:flex;flex-wrap:wrap;gap:8px">`
        +execEntries.map(([ym,v])=>`<div style="background:#f5eeff;border-radius:8px;padding:8px 14px;border-left:3px solid #9B59B6;min-width:110px">`
          +`<div style="font-size:11px;font-weight:700;color:#9B59B6">${{fmtYm(ym)}}</div>`
          +`<div style="font-size:18px;font-weight:800;color:#2c2c2c">${{BRLk(v.valor)}}</div>`
          +`<div style="font-size:10px;color:var(--gray)">${{v.qtd}} proposta${{v.qtd>1?'s':''}}</div>`
          +`</div>`).join('')
        +`</div></div>`;
    }})()
    +`</div>`
    // Bloco Total Geral
  // Renderizar Total Geral no div dedicado (fica acima do filtro)
  const elTG=document.getElementById('total-geral-bloco');
  if(elTG)elTG.innerHTML=
   `<div style="background:linear-gradient(135deg,#076A76 0%,#0a8a99 100%);border-radius:10px;padding:14px 18px;color:#fff;margin-bottom:0">`
   +`<div style="font-size:12px;font-weight:800;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;display:flex;justify-content:space-between;align-items:center">`
   +`<span>Total Geral &mdash; ERP + CRM + Manuais</span>`
   +`<span style="font-size:15px;font-weight:800">${{consolidadoTotQ}} propostas &nbsp;|&nbsp; ${{BRLk(consolidadoTotV)}}</span></div>`
   +mkStatusGrid(consolidadoPip,consolidadoTotQ)
   +`</div>`;

}}

function renderMacro(){{
  const f=getFiltered();
  const pip={{}};
  f.forEach(s=>{{Object.entries(s.pipeline||{{}}).forEach(([st,qtd])=>{{
    if(!pip[st]) pip[st]={{qtd:0,valor:0}};
    pip[st].qtd+=qtd; pip[st].valor+=(s.pipeline_val||{{}})[st]||0;
  }});}});
  const totQtd=Object.values(pip).reduce((a,v)=>a+v.qtd,0);
  const totVal=Object.values(pip).reduce((a,v)=>a+v.valor,0);
  document.getElementById('macro-total').textContent=totQtd;
  document.getElementById('macro-valor').textContent='Valor total: '+BRL(totVal);
  const prelCt=state.prelim?Object.values(state.prelim).flat().length:0;
  document.getElementById('macro-prelim').textContent=prelCt?'incluindo '+prelCt+' proposta(s) preliminar(es)':'';
  const COR_S={{'Aprovada':'#076A76','Enviada':'#41A8B9','Em negociação':'#FBAE4B','Reprovada':'#DC3545','Sem continuidade':'#A9A69F','Cancelada':'#61605B','Preliminar':'#9B59B6'}};
  document.getElementById('macro-status-list').innerHTML=PIP_ORDER.concat(['Preliminar']).filter(k=>pip[k]&&pip[k].qtd>0).map(k=>{{
    const v=pip[k]; const pct=totQtd?Math.round(v.qtd/totQtd*100):0;
    return`<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
      <span style="width:10px;height:10px;border-radius:50%;background:${{COR_S[k]||'#999'}};flex-shrink:0"></span>
      <span style="flex:1;font-size:12px">${{k}}</span>
      <span style="font-size:12px;font-weight:700">${{v.qtd}}</span>
      <span style="font-size:11px;color:var(--gray)">${{pct}}%</span>
      <span style="font-size:11px;color:var(--gray);min-width:70px;text-align:right">${{BRLk(v.valor)}}</span>
    </div>`;
  }}).join('');
  // Gráfico evolução macro
  const meses=Object.keys(DATA.macro).sort();
  killChart('macro');
  charts.macro=new Chart(document.getElementById('cMacro').getContext('2d'),{{
    data:{{labels:meses.map(m=>{{const[y,mo]=m.split('-');return MESES_PT[parseInt(mo)-1]+'/'+y.slice(2)}}),
      datasets:[
        {{type:'bar',label:'Qtd Propostas',data:meses.map(m=>DATA.macro[m].qtd),backgroundColor:'rgba(65,168,185,.6)',borderRadius:3,yAxisID:'y1',order:2}},
        {{type:'line',label:'Valor Total',data:meses.map(m=>DATA.macro[m].valor),borderColor:'var(--accent)',borderWidth:2,pointRadius:2,fill:false,yAxisID:'y2',order:1}},
      ]}},
    options:{{responsive:true,maintainAspectRatio:true,
      plugins:{{legend:{{display:true,position:'top',labels:{{font:{{size:10}}}}}},tooltip:{{callbacks:{{label:c=>c.datasetIndex===0?c.parsed.y+' propostas':BRL(c.parsed.y)}}}}}},
      scales:{{
        y1:{{position:'left',ticks:{{font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}},
        y2:{{position:'right',ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{display:false}}}},
        x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}}
      }}}}
  }});
}}

function renderEntregas(){{
  const E=DATA.entregas;
  // Semana
  const semanas=Object.keys(E.por_semana||{{}}).sort();
  killChart('entSem'); killChart('duracao'); killChart('antec'); killChart('entMes');
  if(semanas.length){{
    charts.entSem=new Chart(document.getElementById('cEntSemana').getContext('2d'),{{
      type:'bar',
      data:{{labels:semanas,datasets:[{{data:semanas.map(s=>E.por_semana[s]),backgroundColor:'rgba(7,106,118,.6)',borderRadius:3}}]}},
      options:{{responsive:true,maintainAspectRatio:true,plugins:{{legend:{{display:false}}}},
        scales:{{x:{{ticks:{{font:{{size:9}},maxRotation:60}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
    }});
  }}
  // Duração média por vertical (Casa vs Eventos)
  const durCasa=E.duracao_casa||{{}};
  const durEvt=E.duracao_eventos||{{}};
  const durAllKeys=Array.from(new Set([...Object.keys(durCasa),...Object.keys(durEvt)])).sort();
  if(durAllKeys.length){{
    const gCasa=E.duracao_media_global&&E.duracao_media_global.Casa;
    const gEvt=E.duracao_media_global&&E.duracao_media_global.Eventos;
    const dg=document.getElementById('dur-global');
    if(dg) dg.innerHTML=(gCasa?'<span style="color:#076A76">■</span> Média do período — Casa: <b>'+gCasa+' dias</b>&nbsp;&nbsp;':'')+
      (gEvt?'<span style="color:#FBAE4B">■</span> Eventos: <b>'+gEvt+' dias</b>':'');
    charts.duracao=new Chart(document.getElementById('cDuracao').getContext('2d'),{{
      type:'bar',
      data:{{labels:durAllKeys.map(m=>{{const[y,mo]=m.split('-');return MESES_PT[parseInt(mo)-1]+'/'+y.slice(2)}}),
        datasets:[
          {{label:'Casa',data:durAllKeys.map(m=>durCasa[m]||null),backgroundColor:'rgba(7,106,118,.75)',borderRadius:3}},
          {{label:'Eventos',data:durAllKeys.map(m=>durEvt[m]||null),backgroundColor:'rgba(251,174,75,.8)',borderRadius:3}},
        ]}},
      options:{{responsive:true,maintainAspectRatio:true,
        plugins:{{legend:{{display:true,labels:{{font:{{size:10}},boxWidth:12}}}},tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+c.parsed.y+' dias'}}}}}},
        scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
    }});
  }}
  // Antecedência
  const antMes=Object.keys(E.antecedencia_media||{{}}).sort();
  if(antMes.length){{
    charts.antec=new Chart(document.getElementById('cAntec').getContext('2d'),{{
      type:'bar',
      data:{{labels:antMes.map(m=>{{const[y,mo]=m.split('-');return MESES_PT[parseInt(mo)-1]+'/'+y.slice(2)}}),
        datasets:[{{data:antMes.map(m=>E.antecedencia_media[m]),backgroundColor:'rgba(251,174,75,.7)',borderRadius:3}}]}},
      options:{{responsive:true,maintainAspectRatio:true,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>c.parsed.y+' dias'}}}}}},
        scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
    }});
  }}

  // Entregas por mês
  const entMes=Object.keys(E.por_mes||{{}}).sort();
  if(entMes.length){{
    charts.entMes=new Chart(document.getElementById('cEntMes').getContext('2d'),{{
      type:'bar',
      data:{{labels:entMes.map(m=>{{const[y,mo]=m.split('-');return MESES_PT[parseInt(mo)-1]+'/'+y.slice(2)}}),
        datasets:[{{data:entMes.map(m=>E.por_mes[m]),backgroundColor:'rgba(176,134,90,.6)',borderRadius:3}}]}},
      options:{{responsive:true,maintainAspectRatio:true,plugins:{{legend:{{display:false}}}},
        scales:{{x:{{ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}}}}}
    }});
  }}
}}

function renderFatFin(){{
  const FF=DATA.fat_fin.por_mes||{{}};
  const f=getFiltered();
  let total=0; const porTipo={{}};
  f.forEach(s=>{{
    const d=FF[s.ym]; if(!d) return;
    total+=d.total;
    Object.entries(d.por_tipo||{{}}).forEach(([t,v])=>{{porTipo[t]=(porTipo[t]||0)+v;}});
  }});
  (state.fatManual||[]).forEach(e=>{{
    const inFilter=f.some(s=>s.ym===e.ym); if(!inFilter) return;
    total+=parseFloat(e.valor||0);
    porTipo[e.tipo_receita]=(porTipo[e.tipo_receita]||0)+parseFloat(e.valor||0);
  }});
  document.getElementById('fin-total').textContent=BRL(total);
  document.getElementById('fin-sub').textContent=total?'Faturas emitidas no período':'Sem dados no período';
  const TIPOS_ORDER=['Casa - PF','Casa - PJ','Eventos - PF','Eventos - PJ','Escritório','Concierge','Outros'];
  const COR_V={{'Casa - PF':'#41A8B9','Casa - PJ':'#076A76','Eventos - PF':'#FBAE4B','Eventos - PJ':'#E07B00','Escritório':'#2E86AB','Concierge':'#9B59B6','Outros':'#A9A69F'}};
  // Por tipo de receita (card)
  const allTipos=[...TIPOS_ORDER,...Object.keys(porTipo).filter(t=>!TIPOS_ORDER.includes(t))];
  document.getElementById('fin-tipos').innerHTML=allTipos.filter(t=>porTipo[t]).map(t=>{{
    const v=porTipo[t]; const pct=total?Math.round(v/total*100):0;
    return`<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
      <span style="width:10px;height:10px;border-radius:50%;background:${{COR_V[t]||'#999'}};flex-shrink:0"></span>
      <span style="flex:1;font-size:12px">${{t}}</span>
      <span style="font-size:12px;font-weight:700">${{BRL(v)}}</span>
      <span style="font-size:11px;color:var(--gray)">${{pct}}%</span>
    </div>`;
  }}).join('')||'<p style="font-size:12px;color:var(--gray)">Sem faturas no período.</p>';
  // ── Gráfico empilhado por categoria ──────────────────────────────────────
  const ymSet=new Set([...Object.keys(FF),...(state.fatManual||[]).map(e=>e.ym)]);
  const meses=[...ymSet].filter(ym=>f.some(s=>s.ym===ym)).sort();
  killChart('fatFin');
  if(!meses.length) return;
  // Montar dados por tipo e mês (ERP)
  const allData={{}};
  meses.forEach(ym=>{{
    const d=FF[ym]||{{total:0,por_tipo:{{}}}};
    allTipos.forEach(t=>{{
      if(!allData[t]) allData[t]={{}};
      allData[t][ym]=(allData[t][ym]||0)+(d.por_tipo[t]||0);
    }});
  }});
  // Adicionar manual ao allData
  (state.fatManual||[]).forEach(e=>{{
    if(!meses.includes(e.ym)) return;
    const t=e.tipo_receita||'Outros';
    if(!allData[t]) allData[t]={{}};
    allData[t][e.ym]=(allData[t][e.ym]||0)+parseFloat(e.valor||0);
  }});
  const barDs=allTipos.filter(t=>meses.some(ym=>allData[t]&&allData[t][ym]>0)).map(t=>
    ({{type:'bar',label:t,data:meses.map(ym=>allData[t]&&allData[t][ym]||0),
      backgroundColor:COR_V[t]||'#999',borderWidth:0,borderRadius:0,stack:'fat',order:2}}));
  charts.fatFin=new Chart(document.getElementById('cFatFin').getContext('2d'),{{
    data:{{
      labels:meses.map(m=>{{const[y,mo]=m.split('-');return MESES_PT[parseInt(mo)-1]+'/'+y.slice(2)}}),
      datasets:[
        ...barDs,
        {{type:'line',label:'Break Even',data:meses.map(ym=>getBreakeven(ym)),
          borderColor:'#DC3545',borderWidth:1.5,borderDash:[5,3],pointRadius:3,fill:false,order:1}},
        {{type:'line',label:'Meta',data:meses.map(ym=>(state.metas&&state.metas[ym])||null),
          borderColor:'#076A76',borderWidth:2,borderDash:[4,2],pointRadius:4,fill:false,order:0,spanGaps:false}},
      ]
    }},
    options:{{responsive:true,maintainAspectRatio:true,
      plugins:{{
        legend:{{position:'bottom',labels:{{font:{{size:11}},boxWidth:12,padding:10,
          filter:item=>item.type!=='line'||item.text==='Break Even'||item.text==='Meta'}}}},
        tooltip:{{callbacks:{{label:c=>` ${{c.dataset.label}}: ${{BRL(c.parsed.y)}}`}}}}
      }},
      scales:{{
        x:{{stacked:true,ticks:{{font:{{size:10}},maxRotation:45}},grid:{{display:false}}}},
        y:{{stacked:true,ticks:{{callback:v=>BRLk(v),font:{{size:10}}}},grid:{{color:'rgba(0,0,0,.04)'}}}}
      }}
    }}
  }});
}}

// ── Propostas Preliminares ───────────────────────────────────────────────────
function loadPrelim(){{try{{state.prelim=JSON.parse(localStorage.getItem('bb_prelim')||'{{}}');}}catch(e){{state.prelim={{}};}}}}
function savePrelimStorage(){{try{{localStorage.setItem('bb_prelim',JSON.stringify(state.prelim));}}catch(e){{}}}}
function openPrelimModal(){{
  loadPrelim(); renderPrelimSaved();
  document.getElementById('prelim-modal').classList.add('open');
}}
function closePrelimModal(){{document.getElementById('prelim-modal').classList.remove('open');renderAll();}}
function _prelimItems(){{
  return Object.entries(state.prelim||{{}}).flatMap(([ym,arr])=>arr.map(e=>({{...e,ym}})));
}}
function savePrelim(){{
  const ym=document.getElementById('prelim-ym').value;
  const vert=document.getElementById('prelim-vert').value;
  const val=parseFloat(document.getElementById('prelim-val').value||0);
  const st=document.getElementById('prelim-status').value;
  const vend=(document.getElementById('prelim-vend')&&document.getElementById('prelim-vend').value)||'';
  const dataExec=(document.getElementById('prelim-data-exec')&&document.getElementById('prelim-data-exec').value)||'';
  if(!ym||!val) return alert('Preencha mês e valor.');
  if(state._editPrelimIdx!=null){{
    const items=_prelimItems();
    items[state._editPrelimIdx]=Object.assign({{}},items[state._editPrelimIdx],{{ym,vertical:vert,valor:val,status:st,vendedor:vend,data_exec:dataExec}});
    state.prelim=items.reduce((acc,x)=>{{if(!acc[x.ym])acc[x.ym]=[];acc[x.ym].push({{vertical:x.vertical,valor:x.valor,status:x.status,vendedor:x.vendedor||'',data_exec:x.data_exec||''}});return acc;}},{{}});
    state._editPrelimIdx=null;
    document.getElementById('prelim-save-btn').textContent='Adicionar';
    document.getElementById('prelim-cancel-edit').style.display='none';
  }} else {{
    if(!state.prelim[ym]) state.prelim[ym]=[];
    state.prelim[ym].push({{vertical:vert,valor:val,status:st,vendedor:vend,data_exec:dataExec}});
  }}
  savePrelimStorage(); renderPrelimSaved();
}}
function editPrelim(idx){{
  const e=_prelimItems()[idx]; if(!e) return;
  document.getElementById('prelim-ym').value=e.ym;
  document.getElementById('prelim-vert').value=e.vertical||'';
  document.getElementById('prelim-val').value=e.valor||'';
  document.getElementById('prelim-status').value=e.status||'';
  if(document.getElementById('prelim-vend')) document.getElementById('prelim-vend').value=e.vendedor||'';
  if(document.getElementById('prelim-data-exec')) document.getElementById('prelim-data-exec').value=e.data_exec||'';
  state._editPrelimIdx=idx;
  document.getElementById('prelim-save-btn').textContent='Salvar alteração';
  document.getElementById('prelim-cancel-edit').style.display='inline-block';
}}
function cancelEditPrelim(){{
  state._editPrelimIdx=null;
  document.getElementById('prelim-save-btn').textContent='Adicionar';
  document.getElementById('prelim-cancel-edit').style.display='none';
}}
function renderPrelimSaved(){{
  const items=_prelimItems();
  document.getElementById('prelim-saved').innerHTML=items.length?
    '<table style="width:100%;font-size:11px;border-collapse:collapse">'
    +'<thead><tr style="background:#f8f9fa"><th style="padding:4px 6px;text-align:left">Mês</th><th>Vertical</th><th>Vendedor</th><th>Status</th><th>Data Execução</th><th>Valor</th><th></th></tr></thead><tbody>'
    +items.map((e,i)=>`<tr style="border-bottom:1px solid #eee">`
      +`<td style="padding:3px 6px">${{e.ym}}</td>`
      +`<td style="padding:3px 6px">${{e.vertical}}</td>`
      +`<td style="padding:3px 6px">${{e.vendedor||'—'}}</td>`
      +`<td style="padding:3px 6px">${{e.status}}</td>`
      +`<td style="padding:3px 6px">${{e.data_exec?e.data_exec.split('-').reverse().join('/'):'—'}}</td>`
      +`<td style="padding:3px 6px">${{BRL(e.valor)}}</td>`
      +`<td style="padding:3px 6px;white-space:nowrap">`
      +`<span style="cursor:pointer;color:var(--primary);margin-right:8px" title="Editar" onclick="editPrelim(${{i}})">✎</span>`
      +`<span style="cursor:pointer;color:var(--red)" title="Remover" onclick="removePrelim(${{i}})">✕</span>`
      +`</td></tr>`).join('')+'</tbody></table>':
    '<p style="font-size:11px;color:var(--gray)">Nenhuma proposta preliminar.</p>';
}}
function removePrelim(idx){{
  state.prelim=_prelimItems().filter((_,i)=>i!==idx).reduce((acc,x)=>{{if(!acc[x.ym])acc[x.ym]=[];acc[x.ym].push({{vertical:x.vertical,valor:x.valor,status:x.status,vendedor:x.vendedor||'',data_exec:x.data_exec||''}});return acc;}},{{}});
  savePrelimStorage(); renderPrelimSaved();
}}
function editPrelimFromList(idx){{
  loadPrelim();
  editPrelim(idx);
  document.getElementById('prelim-modal').classList.add('open');
  renderPrelimSaved();
}}
function renderPrelimList(){{
  loadPrelim();
  const items=_prelimItems();
  document.getElementById('prelim-list').innerHTML=items.length?
    '<table style="width:100%;font-size:12px;border-collapse:collapse">'
    +'<thead><tr style="background:#f8f9fa"><th style="padding:5px 8px;text-align:left">Mês</th>'
    +'<th>Vertical</th><th>Vendedor</th><th>Status</th><th>Data Execução</th><th>Valor</th><th></th></tr></thead><tbody>'
    +items.map((e,i)=>`<tr style="border-bottom:1px solid #eee">`
      +`<td style="padding:4px 8px">${{e.ym}}</td>`
      +`<td style="padding:4px 8px">${{e.vertical}}</td>`
      +`<td style="padding:4px 8px">${{e.vendedor||'—'}}</td>`
      +`<td style="padding:4px 8px">${{e.status}}</td>`
      +`<td style="padding:4px 8px">${{e.data_exec?e.data_exec.split('-').reverse().join('/'):'—'}}</td>`
      +`<td style="padding:4px 8px">${{BRL(e.valor)}}</td>`
      +`<td style="padding:4px 8px;white-space:nowrap">`
      +`<span style="cursor:pointer;color:var(--primary);margin-right:8px" title="Editar" onclick="editPrelimFromList(${{i}})">✎</span>`
      +`<span style="cursor:pointer;color:var(--red)" title="Remover" onclick="removePrelim(${{i}});renderPrelimList();renderPrelimTotalizador()">✕</span>`
      +`</td></tr>`).join('')+'</tbody></table>':
    '<p style="color:var(--gray);font-size:12px">Nenhuma proposta preliminar registrada.</p>';
}}

// ── Faturamento Manual ───────────────────────────────────────────────────────
function loadFatManual(){{try{{state.fatManual=JSON.parse(localStorage.getItem('bb_fat_manual')||'[]');}}catch(e){{state.fatManual=[];}}}}
function saveFatManualStorage(){{try{{localStorage.setItem('bb_fat_manual',JSON.stringify(state.fatManual));}}catch(e){{}}}}
function openFatManualModal(){{
  loadFatManual(); renderFatManualSaved();
  document.getElementById('fat-manual-modal').classList.add('open');
}}
function closeFatManualModal(){{document.getElementById('fat-manual-modal').classList.remove('open');renderFatManualList();renderAll();}}
function saveFatManual(){{
  const ym=document.getElementById('fatm-ym').value;
  const doc=document.getElementById('fatm-doc').value;
  const tipo=document.getElementById('fatm-tipo').value;
  const val=parseFloat(document.getElementById('fatm-val').value||0);
  if(!ym||!val) return alert('Preencha mês e valor.');
  if(!state.fatManual) state.fatManual=[];
  if(state._editFatManualIdx!=null){{
    state.fatManual[state._editFatManualIdx]={{ym,tipo_doc:doc,tipo_receita:tipo,valor:val}};
    state._editFatManualIdx=null;
    document.getElementById('fatm-save-btn').textContent='Adicionar';
    document.getElementById('fatm-cancel-edit').style.display='none';
  }} else {{
    state.fatManual.push({{ym,tipo_doc:doc,tipo_receita:tipo,valor:val}});
  }}
  saveFatManualStorage(); renderFatManualSaved(); renderFatManualList();
}}
function editFatManual(idx){{
  loadFatManual();
  const e=(state.fatManual||[])[idx]; if(!e) return;
  document.getElementById('fatm-ym').value=e.ym;
  document.getElementById('fatm-doc').value=e.tipo_doc||'NFe';
  document.getElementById('fatm-tipo').value=e.tipo_receita||'Outros';
  document.getElementById('fatm-val').value=e.valor||'';
  state._editFatManualIdx=idx;
  document.getElementById('fatm-save-btn').textContent='Salvar alteração';
  document.getElementById('fatm-cancel-edit').style.display='inline-block';
  document.getElementById('fat-manual-modal').classList.add('open');
  renderFatManualSaved();
}}
function cancelEditFatManual(){{
  state._editFatManualIdx=null;
  document.getElementById('fatm-save-btn').textContent='Adicionar';
  document.getElementById('fatm-cancel-edit').style.display='none';
}}
function removeFatManual(idx){{
  if(!state.fatManual) return;
  state.fatManual.splice(idx,1);
  saveFatManualStorage(); renderFatManualSaved(); renderFatManualList(); renderFatFin();
}}
function renderFatManualSaved(){{
  const items=state.fatManual||[];
  document.getElementById('fatm-saved').innerHTML=items.length?
    '<table style="width:100%;font-size:11px;border-collapse:collapse"><thead><tr style="background:#f8f9fa">'
    +'<th style="padding:4px 6px;text-align:left">Mês</th><th style="padding:4px 6px">Doc</th><th style="padding:4px 6px">Tipo</th><th style="padding:4px 6px">Valor</th><th></th></tr></thead><tbody>'+
    items.map((e,i)=>`<tr style="border-bottom:1px solid #eee"><td style="padding:3px 6px">${{e.ym}}</td><td style="padding:3px 6px">${{e.tipo_doc}}</td><td style="padding:3px 6px">${{e.tipo_receita}}</td><td style="padding:3px 6px">${{BRL(e.valor)}}</td>`
    +`<td style="padding:3px 6px;white-space:nowrap"><span style="cursor:pointer;color:var(--primary);margin-right:6px" onclick="editFatManual(${{i}})">&#9998;</span>`
    +`<span style="cursor:pointer;color:var(--red)" onclick="removeFatManual(${{i}})">&#x2715;</span></td></tr>`).join('')+'</tbody></table>':'';
}}
function renderFatManualList(){{
  const items=state.fatManual||[];
  const el=document.getElementById('fat-manual-list');
  if(!el) return;
  el.innerHTML=items.length?
    '<table style="width:100%;font-size:12px;border-collapse:collapse">'
    +'<thead><tr style="background:#f8f9fa">'
    +'<th style="padding:5px 8px;text-align:left;font-weight:600">Mês/Ano</th>'
    +'<th style="padding:5px 8px;text-align:left;font-weight:600">Tipo Documento</th>'
    +'<th style="padding:5px 8px;text-align:left;font-weight:600">Tipo de Receita</th>'
    +'<th style="padding:5px 8px;text-align:right;font-weight:600">Valor</th>'
    +'<th style="padding:5px 8px"></th></tr></thead><tbody>'
    +items.map((e,i)=>`<tr style="border-bottom:1px solid #eee">`
      +`<td style="padding:4px 8px">${{e.ym}}</td>`
      +`<td style="padding:4px 8px">${{e.tipo_doc}}</td>`
      +`<td style="padding:4px 8px">${{e.tipo_receita}}</td>`
      +`<td style="padding:4px 8px;text-align:right;font-weight:600">${{BRL(e.valor)}}</td>`
      +`<td style="padding:4px 8px;white-space:nowrap">`
      +`<span style="cursor:pointer;color:var(--primary);margin-right:8px;font-size:15px" title="Editar" onclick="editFatManual(${{i}})">&#9998;</span>`
      +`<span style="cursor:pointer;color:var(--red);font-size:15px" title="Remover" onclick="removeFatManual(${{i}})">&#x2715;</span>`
      +`</td></tr>`).join('')+'</tbody></table>':
    '<p style="color:var(--gray);font-size:12px;margin:0">Nenhum lançamento manual registrado.</p>';
}}

function toggleCustomPeriod(){{
  const form=document.getElementById('period-custom-form');
  const btn=document.getElementById('btn-period-custom');
  if(form.classList.contains('open')){{clearCustomPeriod();}}
  else {{form.classList.add('open');btn.classList.add('active');}}
}}
function applyCustomPeriod(){{
  const f=document.getElementById('from-ym').value;
  const t=document.getElementById('to-ym').value;
  if(!f||!t||f>t){{alert('Informe um intervalo válido.');return;}}
  state.custom_period=true;state.from_ym=f;state.to_ym=t;
  document.querySelectorAll('.btn-period').forEach(x=>x.classList.remove('active'));
  document.getElementById('sel-month').value='';
  renderAll();
}}
function clearCustomPeriod(){{
  state.custom_period=false;state.from_ym=null;state.to_ym=null;
  document.getElementById('period-custom-form').classList.remove('open');
  document.getElementById('btn-period-custom').classList.remove('active');
  document.getElementById('from-ym').value='';document.getElementById('to-ym').value='';
  renderAll();
}}
document.querySelectorAll('.btn-period').forEach(b=>b.addEventListener('click',()=>{{
  document.querySelectorAll('.btn-period').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');
  const v=b.dataset.year;state.year=v==='all'?'all':parseInt(v);state.month=null;
  state.custom_period=false;state.from_ym=null;state.to_ym=null;
  document.getElementById('period-custom-form').classList.remove('open');
  document.getElementById('btn-period-custom').classList.remove('active');
  document.getElementById('sel-month').value='';renderAll();
}}));
document.getElementById('sel-month').addEventListener('change',e=>{{state.month=e.target.value?parseInt(e.target.value):null;renderAll();}});
document.getElementById('cac-modal').addEventListener('click',e=>{{if(e.target===e.currentTarget)closeCacModal();}});

document.addEventListener('DOMContentLoaded',()=>{{
  loadCacStorage(); loadMetasStorage(); loadBreakevenStorage(); loadPrelim(); loadFatManual(); state.fatManual=state.fatManual||[];
  renderFatManualList();
  const anos=Object.keys(DATA.fat_ano).map(Number);
  const ar=Math.max(...anos);state.year=ar;
  const b=document.querySelector(`.btn-period[data-year="${{ar}}"]`);
  if(b)b.classList.add('active');
  renderAll();
}});
"""

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard Financeiro — Bem Bacana</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>{css}</style>
</head>
<body>
<header class="hdr">
  <div class="hdr-logo">{logo_tag}</div>
  <div class="hdr-right"><strong>Dashboard Financeiro e Comercial</strong><br>Atualizado em {atualizado}</div>
</header>
<div class="filter-bar">
  <span class="filter-label">Período</span>
  <button class="btn-period" data-year="all">Todos</button>
  {anos_opts}
  <select class="month-select" id="sel-month">
    <option value="">Todos os meses</option>
    <option value="1">Janeiro</option><option value="2">Fevereiro</option>
    <option value="3">Março</option><option value="4">Abril</option>
    <option value="5">Maio</option><option value="6">Junho</option>
    <option value="7">Julho</option><option value="8">Agosto</option>
    <option value="9">Setembro</option><option value="10">Outubro</option>
    <option value="11">Novembro</option><option value="12">Dezembro</option>
  </select>
  <button id="btn-period-custom" onclick="toggleCustomPeriod()">&#8596; Per&#237;odo</button>
  <div id="period-custom-form">
    <span style="font-size:11px;color:var(--gray)">De</span>
    <input type="month" class="period-input" id="from-ym">
    <span style="font-size:11px;color:var(--gray)">At&#233;</span>
    <input type="month" class="period-input" id="to-ym">
    <button class="period-apply" onclick="applyCustomPeriod()">Aplicar</button>
    <button class="period-clear" onclick="clearCustomPeriod()">Limpar</button>
  </div>
  <span id="filter-desc"></span>
</div>
<div class="wrap">
  <div class="sec-label" id="kpi-sec-label">Performance</div>
  <div class="grid-kpi">
    <div class="kpi" id="card-fat"></div>
    <div class="kpi" id="card-acum"></div>
    <div class="kpi" id="card-clientes"></div>
    <div class="kpi" id="card-propostas"></div>
    <div class="kpi" id="card-negociacao"></div>
    <div class="kpi" id="card-cac">
      <label>CAC — Custo de Aquisição</label>
      <div class="v" id="cac-value">—</div>
      <div class="sub" id="cac-sub">Clique para informar investimento</div>
      <button class="cac-btn" onclick="openCacModal()">✎ Informar investimento em marketing</button>
    </div>
  </div>
  <!-- ── PIPELINE COMERCIAL — VISÃO GERAL ──────────────────────────────── -->
  <div id="erp-bloco-wrap">
  <div class="sec-label">Comercial Base ERP — Visão Geral{_erp_dt}</div>
  <!-- Bloco 1: Performance Real ERP -->
  <div id="prop-erp-bloco" style="margin-bottom:16px"></div>
  </div>
  <!-- ── COMERCIAL BASE CRM ───────────────────────────────────────────────── -->
  <div class="sec-label" id="sec-crm-label" style="display:none">Comercial Base CRM — Visão Geral{_crm_dt}</div>
  <div id="crm-bloco" style="margin-bottom:16px"></div>
  <!-- ── PREVISÃO DE ENTREGAS CRM ─────────────────────────────────────────── -->
  <div class="sec-label" id="sec-crm-ent-label" style="display:none">Previsão de Entregas CRM</div>
  <div id="crm-entregas" style="margin-bottom:16px"></div>
  <!-- ── PROPOSTAS PRELIMINARES ─────────────────────────────────────────── -->
  <div class="sec-label">Propostas Preliminares (lançamento manual)</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Propostas não registradas no ERP
      <button class="cac-btn" onclick="openPrelimModal()" style="float:right">+ Adicionar</button>
    </div>
    <div id="prelim-list" style="margin-top:8px">
      <p style="color:var(--gray);font-size:12px">Nenhuma proposta preliminar registrada.</p>
    </div>
  </div></div>
  </div><!-- /manual-bloco-wrap -->
  <!-- Resumo/totalizador das preliminares -->
  <div class="manual-bloco-wrap"><div id="prelim-totalizador" style="margin-bottom:8px"></div></div>
  <!-- Propostas Preliminares — Fora do ERP e do CRM (visão geral) -->
  <div id="prop-totalizador" style="margin-bottom:16px"></div>
  <!-- Total Geral — ERP + CRM + Manuais -->
  <div id="total-geral-bloco" style="margin-bottom:14px"></div>
  <!-- Filtro por origem (abaixo do Total Geral) -->
  <div style="display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;align-items:center">
    <span style="font-size:10px;font-weight:700;color:var(--gray);text-transform:uppercase;letter-spacing:1.5px">Filtrar por origem:</span>
    <button class="fonte-btn active" data-fonte="todos" onclick="setFonteFilter('todos',this)">&#9776; Todas as origens</button>
    <button class="fonte-btn" data-fonte="erp"    onclick="setFonteFilter('erp',this)">&#9654; Somente ERP</button>
    <button class="fonte-btn" data-fonte="crm"    onclick="setFonteFilter('crm',this)">&#9654; Somente CRM</button>
    <button class="fonte-btn" data-fonte="manual" onclick="setFonteFilter('manual',this)">&#9654; Somente Manual</button>
  </div>
  <!-- Card de contexto % (visível apenas quando fonte isolada) -->
  <div id="fonte-context-card" style="display:none"></div>
  <!-- ── EVOLUÇÃO DE CLIENTES ──────────────────────────────────────────── -->
  <div class="sec-label">Evolução de Clientes — Base ERP</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Novos vs Antigos por Mês (ERP)</div>
    <div id="evol-cli-erp" style="margin-top:8px"><p style="color:var(--gray);font-size:12px">Carregando...</p></div>
  </div></div>
  <div class="sec-label">Evolução de Clientes — Base CRM</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Novos vs Antigos por Mês (CRM) <span style="font-size:10px;font-weight:400;color:var(--gray)">— Antigo = já apareceu no ERP em qualquer época</span></div>
    <div id="evol-cli-crm" style="margin-top:8px"><p style="color:var(--gray);font-size:12px">Carregando...</p></div>
  </div></div>
  <div class="sec-label">Evolução de Clientes — Base Lançamento Manual</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Novos vs Antigos por Mês (Lançamentos Manuais)</div>
    <div id="evol-cli-man" style="margin-top:8px"><p style="color:var(--gray);font-size:12px">Carregando...</p></div>
  </div></div>

  <div class="sec-label">Evolução de Faturamento — ERP (only)</div>
  <div class="row2">
    <div class="card">
      <div class="ctitle">Faturamento Mensal vs Break Even / Meta <button class="cac-btn" onclick="openMetasModal()" style="float:right">&#9998; Metas</button> <button class="cac-btn" onclick="openBreakevenModal()" style="float:right;margin-right:6px">&#9998; Break Even</button></div>
      <canvas id="cFat" height="95"></canvas>
    </div>
    <div class="card">
      <div class="ctitle">Status de Propostas</div>
      <canvas id="cPip" height="140"></canvas>
      <div id="pip-legend" style="margin-top:12px"></div>
    </div>
  </div>
  <div class="sec-label">Faturamento por Vertical</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Composição por Vertical (R$)
      <select id="sel-n-meses" onchange="renderVertChart()">
        <option value="6">Últimos 6 meses</option>
        <option value="12" selected>Últimos 12 meses</option>
        <option value="24">Últimos 24 meses</option>
        <option value="0">Todos</option>
      </select>
    </div>
    <canvas id="cVert" height="55"></canvas>
  </div></div>
  <div class="sec-label">Mix de Receita — Locação vs Serviço/Frete</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Locação vs Serviço/Frete por Mês</div>
    <div id="mix-content"></div>
  </div></div>
  <div class="sec-label">Análise Detalhada</div>
  <div class="row3">
    <div class="card">
      <div class="ctitle">Ticket Médio por Vertical
        <select id="sel-ticket" onchange="renderTicket()">
          <option value="all">Acumulado</option>
          <option value="year">Ano selecionado</option>
          <option value="month">Mês selecionado</option>
        </select>
      </div>
      <table><thead><tr><th>Vertical</th><th>Contratos</th><th>Ticket Médio</th><th>Total</th></tr></thead>
      <tbody id="ticket-tbody"></tbody></table>
    </div>
    <div class="card">
      <div class="ctitle">Top 10 Clientes — Receita
        <select id="sel-top" onchange="renderTop()">
          <option value="all">Acumulado</option>
          <option value="year">Ano selecionado</option>
          <option value="month">Mês selecionado</option>
        </select>
      </div>
      <canvas id="cTop" height="220"></canvas>
    </div>
    <div class="card">
      <div class="ctitle">Faturamento por Ano</div>
      <div id="fat-anos"></div>
    </div>
  </div>
  <div class="sec-label">Frequência de Locação — Clientes Antigos (já alugaram antes)</div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Ranking de Clientes
      <button class="cac-btn" id="freq-sort-btn" onclick="toggleFreqSort()" style="float:right">Ordenar por Receita</button>
    </div>
    <table><thead><tr><th style="text-align:center;width:28px">#</th><th>Cliente</th><th>Vertical</th><th id="freq-sort-col" style="text-align:right">Locações</th><th style="text-align:right">Receita Total</th><th style="text-align:right">Ticket Médio</th></tr></thead>
    <tbody id="freq-tbody"></tbody></table>
  </div></div>
</div>

  <!-- ── MACRO PROPOSTAS ───────────────────────────────────────────────── -->
  <div class="sec-label" id="macro-sec-label">Visão Macro — Todas as Propostas</div>
  <div class="row3">
    <div class="card">
      <div class="ctitle">Total de Propostas no Período</div>
      <div id="macro-total" class="v" style="font-size:28px">—</div>
      <div class="sub" id="macro-valor">—</div>
      <div class="sub2" id="macro-prelim" style="margin-top:4px"></div>
    </div>
    <div class="card" style="grid-column:span 2">
      <div class="ctitle">Propostas por Status — Período Selecionado</div>
      <div id="macro-status-list" style="margin-top:8px"></div>
    </div>
  </div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Evolução de Propostas por Mês</div>
    <canvas id="cMacro" height="80"></canvas>
  </div></div>

  <!-- ── ANÁLISE DE ENTREGAS ───────────────────────────────────────────── -->
  <div class="sec-label">Análise de Entregas e Retiradas</div>
  <div class="row3">
    <div class="card">
      <div class="ctitle">Concentração de Entregas — por Semana</div>
      <canvas id="cEntSemana" height="160"></canvas>
    </div>
    <div class="card">
      <div class="ctitle">Duração Média das Locações por Vertical (dias)</div>
      <div id="dur-global" style="font-size:11px;color:var(--gray);margin-bottom:6px;padding:3px 0"></div>
      <canvas id="cDuracao" height="145"></canvas>
    </div>
    <div class="card">
      <div class="ctitle">Antecedência Média (dias entre proposta e entrega)</div>
      <canvas id="cAntec" height="160"></canvas>
    </div>
  </div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Entregas por Mês</div>
    <canvas id="cEntMes" height="80"></canvas>
  </div></div>

  <!-- ── FATURAS EMITIDAS ───────────────────────────────────────────────── -->
  <div class="sec-label">Faturas Emitidas — Financeiro{_fat_dt}</div>
  <div class="row3">
    <div class="card">
      <div class="ctitle">Total Faturado (faturas emitidas)</div>
      <div id="fin-total" class="v" style="font-size:26px">—</div>
      <div class="sub" id="fin-sub">Período selecionado</div>
    </div>
    <div class="card" style="grid-column:span 2">
      <div class="ctitle">Por Tipo de Receita
        <button class="cac-btn" onclick="openFatManualModal()" style="float:right">+ Lançamento Manual</button>
      </div>
      <div id="fin-tipos" style="margin-top:8px"></div>
    </div>
  </div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Resumo de Faturas Lançadas Manualmente</div>
    <div id="fat-manual-list" style="margin-top:8px"></div>
  </div></div>
  <div class="row-full"><div class="card">
    <div class="ctitle">Evolução de Faturas Emitidas por Mês</div>
    <canvas id="cFatFin" height="80"></canvas>
  </div></div>

  <!-- ── INADIMPLÊNCIA ──────────────────────────────────────────────────── -->
  <div class="sec-label" style="display:flex;justify-content:space-between;align-items:center">
    <span>Controle de Inadimplência</span>
    <div style="display:flex;gap:8px">
      <button class="cac-btn" onclick="openInadimModal()">+ Registrar</button>
    </div>
  </div>
  <div id="inadim-section" style="margin-bottom:16px"></div>

<div class="footer">Dashboard Financeiro · Bem Bacana Locações · {atualizado}</div>
<div class="modal-backdrop" id="cac-modal">
  <div class="modal">
    <h3>Investimento em Marketing</h3>
    <p class="desc">CAC = Investimento ÷ Clientes Novos. Valores salvos no navegador.</p>
    <div id="cac-inputs"></div>
    <button class="btn-primary" onclick="saveCac()">Salvar e Calcular</button>
    <button class="btn-sec" onclick="closeCacModal()">Fechar</button>
  </div>
</div>
<div class="modal-backdrop" id="prelim-modal">
  <div class="modal" style="max-width:480px">
    <h3>Proposta Preliminar</h3>
    <p class="desc">Propostas elaboradas fora do ERP. Entram como "Preliminar" no pipeline.</p>
    <div style="display:grid;gap:10px;margin-top:12px">
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">MÊS/ANO</label>
        <input id="prelim-ym" type="month" class="modal-input"></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">VERTICAL</label>
        <select id="prelim-vert" class="modal-input">
          <option>Eventos</option><option>Casa</option><option>Escritório</option>
          <option>Concierge</option><option>Seminovos</option>
        </select></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">VALOR (R$)</label>
        <input id="prelim-val" type="number" class="modal-input" placeholder="0.00"></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">VENDEDOR / CONSULTOR</label>
        <select id="prelim-vend" class="modal-input">
          <option value="">— Não informado —</option>
          <option>Atendimento 1</option>
          <option>Atendimento 2</option>
          <option>Atendimento 3</option>
          <option>Atendimento 4</option>
          <option>Ana Lu 7899</option>
          <option>Ana Lu 2277</option>
          <option>Diretoria</option>
          <option>Eventos</option>
          <option>Casa</option>
          <option>Escritório</option>
          <option>Concierge</option>
          <option>Seminovos</option>
        </select></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">STATUS</label>
        <select id="prelim-status" class="modal-input">
          <option value="Enviada">Enviada</option>
          <option value="Em negociação">Em negociação</option>
          <option value="Aprovada">Aprovada</option>
          <option value="Reprovada">Reprovada</option>
          <option value="Cancelada">Cancelada</option>
          <option value="Sem continuidade">Cliente não deu continuidade</option>
          <option value="Preliminar">Preliminar (genérico)</option>
        </select></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">DATA DE EXECUÇÃO (previsão de início)</label>
        <input id="prelim-data-exec" type="date" class="modal-input"></div>
    </div>
    <div style="margin-top:14px;display:flex;gap:8px;flex-wrap:wrap">
      <button class="btn-primary" id="prelim-save-btn" onclick="savePrelim()">Adicionar</button>
      <button id="prelim-cancel-edit" class="btn-sec" style="background:#888;color:#fff;display:none" onclick="cancelEditPrelim()">Cancelar edição</button>
      <button class="btn-sec" onclick="closePrelimModal()">Fechar</button>
    </div>
    <div id="prelim-saved" style="margin-top:12px;max-height:160px;overflow-y:auto"></div>
  </div>
</div>
<div class="modal-backdrop" id="fat-manual-modal">
  <div class="modal" style="max-width:480px">
    <h3>Lançamento Manual — Fatura</h3>
    <p class="desc">Faturas emitidas não constantes na exportação do sistema.</p>
    <div style="display:grid;gap:10px;margin-top:12px">
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">MÊS/ANO</label>
        <input id="fatm-ym" type="month" class="modal-input"></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">TIPO DOCUMENTO</label>
        <select id="fatm-doc" class="modal-input">
          <option value="NFe">NFe</option>
          <option value="DANFe">DANFe</option>
          <option value="Recibo">Recibo de Locação</option>
        </select></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">TIPO RECEITA</label>
        <select id="fatm-tipo" class="modal-input">
          <option>Casa - PF</option><option>Casa - PJ</option>
          <option>Eventos - PF</option><option>Eventos - PJ</option>
          <option>Escritório</option><option>Concierge</option><option>Outros</option>
        </select></div>
      <div><label style="font-size:11px;font-weight:700;color:var(--gray)">VALOR (R$)</label>
        <input id="fatm-val" type="number" class="modal-input" placeholder="0.00"></div>
    </div>
    <div style="margin-top:14px;display:flex;gap:8px;flex-wrap:wrap">
      <button class="btn-primary" id="fatm-save-btn" onclick="saveFatManual()">Adicionar</button>
      <button id="fatm-cancel-edit" class="btn-sec" style="background:#888;color:#fff;display:none" onclick="cancelEditFatManual()">Cancelar edição</button>
      <button class="btn-sec" onclick="closeFatManualModal()">Fechar</button>
    </div>
    <div id="fatm-saved" style="margin-top:12px;max-height:160px;overflow-y:auto"></div>
  </div>
</div>
<div class="modal-backdrop" id="metas-modal" style="display:none">
  <div class="modal" style="max-width:400px">
    <h3>Metas Mensais</h3>
    <p class="desc">Break even fixo: R$ 150.000/mês. Defina metas específicas opcionais por mês (salvas no navegador).</p>
    <div id="metas-inputs" style="max-height:300px;overflow-y:auto;margin-top:12px"></div>
    <div style="margin-top:14px;display:flex;gap:8px">
      <button class="btn-primary" onclick="saveMetas()">Salvar</button>
      <button class="btn-sec" onclick="closeMetasModal()">Fechar</button>
    </div>
  </div>
</div>
<script>{js}</script>

<!-- ── MODAL INADIMPLÊNCIA ─────────────────────────────────────── -->
<div id="inadim-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9999;align-items:center;justify-content:center">
  <div style="background:#fff;border-radius:12px;padding:28px;max-width:520px;width:90%;box-shadow:0 8px 40px rgba(0,0,0,.25)">
    <div style="font-size:14px;font-weight:800;color:#DC3545;text-transform:uppercase;letter-spacing:1px;margin-bottom:18px;display:flex;justify-content:space-between;align-items:center">
      Registrar Inadimplência
      <button onclick="closeInadimModal()" style="background:none;border:none;font-size:20px;cursor:pointer;color:var(--gray)">×</button>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div>
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Mês/Ano *</label>
        <input type="month" id="inadim-ym" style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Vertical *</label>
        <select id="inadim-vertical" style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
          <option value="">Selecione...</option>
          <option>Eventos</option><option>Casa</option><option>Escritório</option>
          <option>Concierge</option><option>Seminovos</option>
        </select>
      </div>
      <div style="grid-column:1/-1">
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Razão Social *</label>
        <input type="text" id="inadim-razao" placeholder="Nome do cliente..." style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Nº Contrato</label>
        <input type="text" id="inadim-contrato" placeholder="2026-XXX" style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
      </div>
      <div>
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Valor (R$) *</label>
        <input type="number" id="inadim-valor" placeholder="0,00" step="0.01" style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
      </div>
      <div style="grid-column:1/-1">
        <label style="font-size:11px;font-weight:700;color:var(--gray);display:block;margin-bottom:4px">Observações</label>
        <input type="text" id="inadim-obs" placeholder="Parcela X/Y, vencimento, tentativas de contato..." style="width:100%;padding:8px;border:1px solid var(--border);border-radius:6px;font-size:13px">
      </div>
    </div>
    <div style="margin-top:18px;display:flex;gap:10px;justify-content:flex-end">
      <button onclick="closeInadimModal()" style="padding:8px 20px;border:1px solid var(--border);border-radius:6px;background:#f0f0f0;cursor:pointer;font-weight:600">Cancelar</button>
      <button onclick="saveInadim()" style="padding:8px 24px;border:none;border-radius:6px;background:#DC3545;color:#fff;cursor:pointer;font-weight:700">Salvar</button>
    </div>
  </div>
</div>

</body>
</html>"""
    return html


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def _sort_key_arquivo(nome):
    """
    Chave de ordenação para selecionar o arquivo mais recente entre candidatos.
    Suporta dois formatos no nome:
      - (DD.MM.AAAA) → converte para YYYYMMDD (retorna string tipo '20260509')
      - (N) número sequencial → '0000000NNN' (menor que qualquer data)
    Arquivo sem marcador retorna '0000000000'.
    """
    import re
    m = re.search(r'\((\d{2})\.(\d{2})\.(\d{4})\)', nome)
    if m:
        return f"{m.group(3)}{m.group(2)}{m.group(1)}"
    m2 = re.search(r'\((\d+)\)', nome)
    if m2:
        return f"0000{int(m2.group(1)):06d}"
    return '0000000000'



def _sort_key_arquivo(nome):
    """Chave de ordenação para selecionar o arquivo mais recente (data > número sequencial)."""
    import re
    m = re.search(r'\((\d{2})\.(\d{2})\.(\d{4})\)', nome)
    if m:
        return f"{m.group(3)}{m.group(2)}{m.group(1)}"  # YYYYMMDD
    m2 = re.search(r'\((\d+)\)', nome)
    if m2:
        return f"0000{int(m2.group(1)):06d}"
    return '0000000000'


def main():
    import re as _re

    base = os.path.dirname(os.path.abspath(__file__))
    proj = os.path.join(base, '..')
    dados  = os.path.join(proj, 'dados')
    output = os.path.join(proj, 'output')
    os.makedirs(output, exist_ok=True)

    xls_files = [f for f in os.listdir(dados) if f.lower().endswith(('.xlsx','.xls'))]

    # ── Planilha histórica (Dash BB) ─────────────────────────────────────────
    hist_file = next((os.path.join(dados,f) for f in xls_files if 'dash' in f.lower()), None)

    # ── ERP / Eloca: aceita "Relatorio*" E novo "Planilh* Propostas ERP (DD.MM.AAAA)" ──
    erp_cands = [f for f in xls_files if
        'relat' in f.lower() or
        ('planil' in f.lower() and 'proposta' in f.lower() and 'erp' in f.lower())]
    erp_cands.sort(key=_sort_key_arquivo, reverse=True)
    eloca_file = os.path.join(dados, erp_cands[0]) if erp_cands else None

    # ── Faturamento: aceita "ftp*" E novo "Planilha de Faturamento (DD.MM.AAAA)" ──
    fat_cands = [f for f in xls_files if
        'ftp' in f.lower() or
        ('planil' in f.lower() and 'faturamento' in f.lower() and 'proposta' not in f.lower())]
    fat_cands.sort(key=_sort_key_arquivo, reverse=True)
    fat_file = os.path.join(dados, fat_cands[0]) if fat_cands else None

    # ── CRM: aceita "Preliminar* (sem CRM no nome)" E novo "Planilh* CRM (DD.MM.AAAA)" ──
    crm_cands = [f for f in xls_files if
        ('preliminar' in f.lower() and 'crm' not in f.lower()) or
        ('planil' in f.lower() and 'crm' in f.lower())]
    crm_cands.sort(key=_sort_key_arquivo, reverse=True)
    crm_file = os.path.join(dados, crm_cands[0]) if crm_cands else None

    if not hist_file:  sys.exit('Arquivo historico nao encontrado em dados/')
    if not eloca_file: sys.exit('Arquivo Eloca nao encontrado em dados/')

    # ── Extrair data de atualização do nome do arquivo ───────────────────────
    def _data_arquivo(nome):
        m = _re.search(r'\((\d{2})\.(\d{2})\.(\d{4})\)', nome)
        if m: return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
        m2 = _re.search(r'(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})', nome)
        if m2:
            aa = m2.group(3); aa = ('20'+aa) if len(aa)==2 else aa
            return f"{int(m2.group(1)):02d}/{int(m2.group(2)):02d}/{aa}"
        return None

    logo_b64 = carregar_logo(proj)
    manual   = ler_manual_input(dados)

    print(f'[1/5] Planilha2  <- {os.path.basename(hist_file)}')
    p2 = ler_planilha2(hist_file)
    print(f'      {len(p2)} registros (ate dez/2025)')
    print(f'[2/5] ERP_CRM    <- {os.path.basename(hist_file)}')
    erp = ler_erp_crm(hist_file)
    print(f'      {len(erp)} registros')
    print(f'[3/5] Plan CAC   <- {os.path.basename(hist_file)}')
    cac = ler_cac(hist_file)
    print(f'      {len(cac)} meses')
    print(f'[4/5] Eloca      <- {os.path.basename(eloca_file)}')
    eloca = ler_eloca(eloca_file)
    print(f'      {len(eloca)} registros')

    faturamento = []
    if fat_file:
        print(f'[4b]  Faturamento <- {os.path.basename(fat_file)}')
        faturamento = ler_faturamento(fat_file)
        print(f'      {len(faturamento)} faturas')

    crm = []
    if crm_file:
        print(f'[4c]  CRM Prelim. <- {os.path.basename(crm_file)}')
        crm = ler_crm(crm_file)
        nao_erp = sum(1 for r in crm if not r['em_erp'])
        print(f'      {len(crm)} registros ({nao_erp} fora do ERP)')

    fonte_datas = {
        'erp':         _data_arquivo(os.path.basename(eloca_file)) or '',
        'faturamento': _data_arquivo(os.path.basename(fat_file))   or '' if fat_file  else '',
        'crm':         _data_arquivo(os.path.basename(crm_file))   or '' if crm_file  else '',
    }

    print('[5/5] Calculando KPIs e gerando HTML...')
    kpis = calcular(p2, eloca, erp, cac, faturamento=faturamento, manual=manual, crm=crm)
    kpis['fonte_datas'] = fonte_datas
    html = gerar_html(kpis, logo_b64)

    out_path = os.path.join(output, 'dashboard.html')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'Dashboard salvo em: {out_path}')

if __name__ == '__main__':
    main()
